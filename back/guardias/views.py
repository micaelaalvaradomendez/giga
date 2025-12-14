"""
Views para la app guardias - Implementacion de Fase 1 con logica existente.
"""

from rest_framework import status
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.utils import timezone
from django.db.models import Q, Sum, Count, F
from django.db import transaction
from datetime import datetime, date
from django.contrib.staticfiles import finders
import logging

from .models import Cronograma, Guardia, ResumenGuardiaMes, ReglaPlus, ParametrosArea, Feriado, HoraCompensacion
# Importar funciones de validacion de Dias laborables
from asistencia.views import es_dia_laborable, get_motivo_no_laborable
from auditoria.models import Auditoria
from .serializers import (
    CronogramaExtendidoSerializer, GuardiaResumenSerializer,
    ResumenGuardiaMesExtendidoSerializer, ReglaPlusSerializer,
    ParametrosAreaSerializer, FeriadoSerializer,
    PlanificacionCronogramaSerializer, CalculoPlusSerializer, AprobacionPlusSerializer,
    HoraCompensacionSerializer, CrearCompensacionSerializer, AprobacionCompensacionSerializer, ResumenCompensacionSerializer
)
from .utils import CalculadoraPlus, PlanificadorCronograma, ValidadorHorarios
from .services.reportes import obtener_datos_reporte, ReporteError
from .services.reportes import obtener_datos_reporte, ReporteError

# RBAC Permissions
from common.permissions import (
    IsAuthenticatedGIGA, obtener_agente_sesion, obtener_rol_agente,
    obtener_areas_jerarquia
)

logger = logging.getLogger(__name__)


class ReglaPlusViewSet(viewsets.ModelViewSet):
    """ViewSet para gestiÃ³n de reglas de plus salarial"""

    queryset = ReglaPlus.objects.all()
    serializer_class = ReglaPlusSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtros opcionales
        activa = self.request.query_params.get('activa')
        vigente = self.request.query_params.get('vigente')

        if activa is not None:
            queryset = queryset.filter(activa=activa.lower() == 'true')

        if vigente is not None and vigente.lower() == 'true':
            hoy = date.today()
            queryset = queryset.filter(
                activa=True,
                vigente_desde__lte=hoy
            ).filter(
                Q(vigente_hasta__isnull=True) | Q(vigente_hasta__gte=hoy)
            )

        return queryset.order_by('-porcentaje_plus', 'nombre')

    @action(detail=True, methods=['post'])
    def simular(self, request, pk=None):
        """Simula la aplicaciÃ³n de una regla de plus"""
        regla = self.get_object()

        horas_test = request.data.get('horas_efectivas', 160)

        resultado = {
            'regla': regla.nombre,
            'horas_minimas_requeridas': regla.horas_minimas_mensuales,
            'horas_testeo': horas_test,
            'plus_aplicable': float(regla.porcentaje_plus) if horas_test >= regla.horas_minimas_mensuales else 0,
            'aplica': horas_test >= regla.horas_minimas_mensuales
        }

        return Response(resultado)


class ParametrosAreaViewSet(viewsets.ModelViewSet):
    """ViewSet para parÃ¡metros de control horario por Ã¡rea"""

    queryset = ParametrosArea.objects.all()
    serializer_class = ParametrosAreaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtrar por Ã¡rea
        area_id = self.data_in.get('area')
        if area_id:
            queryset = queryset.filter(id_area=area_id)

        # Filtrar solo vigentes
        vigentes = self.request.query_params.get('vigentes')
        if vigentes and vigentes.lower() == 'true':
            hoy = date.today()
            queryset = queryset.filter(
                activo=True,
                vigente_desde__lte=hoy
            ).filter(
                Q(vigente_hasta__isnull=True) | Q(vigente_hasta__gte=hoy)
            )

        return queryset.order_by('-vigente_desde')


class FeriadoViewSet(viewsets.ModelViewSet):
    """ViewSet para gestiÃ³n de feriados con soporte multi-Dia"""

    queryset = Feriado.objects.all()
    serializer_class = FeriadoSerializer
    permission_classes = []  # Temporalmente sin autenticaciÃ³n para debugging

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtros por anio - ahora busca en rango de fechas
        anio = self.request.query_params.get('anio')
        if anio:
            queryset = queryset.filter(
                Q(fecha_inicio__year=anio) |
                Q(fecha_fin__year=anio)
            )

        # Filtros por tipo
        tipo = self.request.query_params.get('tipo')
        if tipo == 'nacional':
            queryset = queryset.filter(es_nacional=True)
        elif tipo == 'provincial':
            queryset = queryset.filter(es_provincial=True)
        elif tipo == 'local':
            queryset = queryset.filter(es_local=True)

        return queryset.filter(activo=True).order_by('fecha_inicio', 'fecha_fin')

    def create(self, request, *args, **kwargs):
        """Override del mÃ©todo create para manejar repeticiÃ³n anual"""
        repetir_anualmente = request.data.get('repetir_anualmente', False)

        if not repetir_anualmente:
            # Comportamiento normal
            return super().create(request, *args, **kwargs)

        # Crear feriado con repeticiÃ³n anual (5 anios hacia adelante)
        try:

            fecha_inicio_str = request.data.get('fecha_inicio')
            fecha_fin_str = request.data.get('fecha_fin', fecha_inicio_str)

            fecha_inicio_base = datetime.strptime(
                fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin_base = datetime.strptime(
                fecha_fin_str, '%Y-%m-%d').date()

            feriados_creados = []
            anios_creados = []

            with transaction.atomic():
                # Crear feriado para los prÃ³ximos 5 anios
                for i in range(5):
                    # Calcular las fechas para este anio
                    fecha_inicio_anio = fecha_inicio_base.replace(
                        year=fecha_inicio_base.year + i)
                    fecha_fin_anio = fecha_fin_base.replace(
                        year=fecha_fin_base.year + i)

                    # Verificar si ya existe un feriado con el mismo nombre en esas fechas
                    existe = Feriado.objects.filter(
                        nombre=request.data.get('nombre'),
                        fecha_inicio=fecha_inicio_anio,
                        fecha_fin=fecha_fin_anio
                    ).exists()

                    if not existe:
                        feriado_data = request.data.copy()
                        feriado_data['fecha_inicio'] = fecha_inicio_anio
                        feriado_data['fecha_fin'] = fecha_fin_anio

                        serializer = self.get_serializer(data=feriado_data)
                        serializer.is_valid(raise_exception=True)
                        feriado = serializer.save()

                        feriados_creados.append(
                            FeriadoSerializer(feriado).data)
                        anios_creados.append(str(fecha_inicio_anio.year))

            return Response({
                'success': True,
                'message': f'Se crearon {len(feriados_creados)} feriados para los prÃ³ximos 5 anios',
                'data': {
                    'feriados_creados': len(feriados_creados),
                    'anios': anios_creados,
                    'feriados': feriados_creados
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                'error': f'Error creando feriados anuales: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def verificar_fecha(self, request):
        """Verifica si una fecha especÃ­fica es feriado - ahora soporta mÃºltiples feriados por fecha"""
        fecha_str = request.data.get('fecha')
        if not fecha_str:
            return Response(
                {'error': 'Fecha requerida'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from datetime import datetime
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()

            # Obtener todos los feriados que incluyen esta fecha
            feriados = Feriado.feriados_en_fecha(fecha)
            es_feriado = feriados.exists()

            feriados_info = []
            if es_feriado:
                feriados_info = FeriadoSerializer(feriados, many=True).data

            return Response({
                'fecha': fecha_str,
                'es_feriado': es_feriado,
                'cantidad_feriados': feriados.count(),
                'feriados': feriados_info
            })

        except ValueError:
            return Response(
                {'error': 'Formato de fecha invÃ¡lido. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get', 'post'])
    def por_mes(self, request):
        """Obtiene feriados para un mes especifico (optimizado para calendario)"""
        anio = request.query_params.get('anio')
        mes = request.query_params.get('mes')

        if not anio or not mes:
            return Response(
                {'error': 'anio y mes son requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Fechas del mes
            from datetime import date, timedelta
            primer_dia = date(int(anio), int(mes), 1)
            if int(mes) == 12:
                ultimo_dia = date(int(anio) + 1, 1, 1) - timedelta(days=1)
            else:
                ultimo_dia = date(int(anio), int(mes) + 1,
                                  1) - timedelta(days=1)

            # Feriados que intersectan con el mes
            feriados = Feriado.feriados_en_rango(primer_dia, ultimo_dia)

            return Response({
                'anio': anio,
                'mes': mes,
                'primer_dia': primer_dia,
                'ultimo_dia': ultimo_dia,
                'total_feriados': feriados.count(),
                'feriados': FeriadoSerializer(feriados, many=True).data
            })

        except ValueError as e:
            return Response(
                {'error': f'anio o mes invÃ¡lidos: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['post'])
    def verificar_rango(self, request):
        """Verifica feriados en un rango de fechas"""
        fecha_inicio = request.data.get('fecha_inicio')
        fecha_fin = request.data.get('fecha_fin')

        if not fecha_inicio or not fecha_fin:
            return Response(
                {'error': 'Fecha inicio y fin requeridas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from datetime import datetime
            fecha_inicio_obj = datetime.strptime(
                fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

            if fecha_fin_obj < fecha_inicio_obj:
                return Response(
                    {'error': 'La fecha fin debe ser mayor o igual a la fecha inicio'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            feriados = Feriado.feriados_en_rango(
                fecha_inicio_obj, fecha_fin_obj)

            return Response({
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'dias_consultados': (fecha_fin_obj - fecha_inicio_obj).days + 1,
                'total_feriados': feriados.count(),
                'feriados': FeriadoSerializer(feriados, many=True).data
            })

        except ValueError:
            return Response(
                {'error': 'Formato de fecha invÃ¡lido. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )

    def perform_create(self, serializer):
        """Override para agregar auditorÃ­a en creaciÃ³n"""
        feriado = serializer.save()
        self._crear_auditoria_feriado(
            'CREAR', feriado.id_feriado, None, serializer.data)

    def perform_update(self, serializer):
        """Override para agregar auditorÃ­a en actualizaciÃ³n"""
        feriado_anterior = self.get_object()
        valor_previo = FeriadoSerializer(feriado_anterior).data

        feriado = serializer.save()
        self._crear_auditoria_feriado(
            'MODIFICAR', feriado.id_feriado, valor_previo, serializer.data)

    def perform_destroy(self, instance):
        """Override para agregar auditorÃ­a en eliminaciÃ³n"""
        valor_previo = FeriadoSerializer(instance).data
        self._crear_auditoria_feriado(
            'ELIMINAR', instance.id_feriado, valor_previo, None)
        instance.delete()

    def _crear_auditoria_feriado(self, accion, feriado_id, valor_previo=None, valor_nuevo=None):
        """Crear registro de auditorÃ­a para cambios en feriados"""
        try:
            from auditoria.models import Auditoria
            from django.utils import timezone
            import logging
            logger = logging.getLogger(__name__)

            # Obtener el agente del usuario autenticado
            agente_id = None
            if hasattr(self.request.user, 'agente'):
                agente_id = self.request.user.agente.id

            Auditoria.objects.create(
                pk_afectada=feriado_id,
                nombre_tabla='feriado',
                creado_en=timezone.now(),
                valor_previo=valor_previo,
                valor_nuevo=valor_nuevo,
                accion=accion,
                id_agente_id=agente_id
            )
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al crear auditorÃ­a de feriado: {str(e)}")


class CronogramaViewSet(viewsets.ModelViewSet):
    """ViewSet extendido para cronogramas con planificaciÃ³n automÃ¡tica"""

    queryset = Cronograma.objects.all()
    serializer_class = CronogramaExtendidoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Optimizar consultas con select_related
        queryset = queryset.select_related(
            'id_area', 'id_jefe', 'id_director',
            'creado_por_id', 'aprobado_por_id'
        )

        # Filtros
        area_id = self.data_in.get('area')
        estado = self.request.query_params.get('estado')
        tipo = self.request.query_params.get('tipo')

        if area_id:
            queryset = queryset.filter(id_area=area_id)
        if estado:
            queryset = queryset.filter(estado=estado)
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        return queryset.order_by('-creado_en')

    @action(detail=False, methods=['post'])
    def crear_con_guardias(self, request):
        """Crea un cronograma con mÃºltiples guardias y registra en auditorÃ­a"""
        try:
            data = request.data

            # Debug: Log de datos recibidos
            print(f"ðŸ“‹ Datos recibidos en crear_con_guardias: {data}")

            # Validar datos requeridos
            required_fields = ['nombre', 'tipo', 'id_area',
                               'fecha', 'hora_inicio', 'hora_fin', 'agentes']
            for field in required_fields:
                if field not in data:
                    return Response(
                        {'error': f'Campo requerido: {field}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Validar que haya agentes
            if not data['agentes'] or len(data['agentes']) == 0:
                return Response(
                    {'error': 'Debe seleccionar al menos un agente'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar que la fecha sea fin de semana o feriado
            try:
                from datetime import datetime as dt
                from asistencia.views import es_dia_laborable, get_motivo_no_laborable

                fecha_guardia = dt.strptime(data['fecha'], '%Y-%m-%d').date()
                logger.info(f"Validando fecha para guardia: {fecha_guardia}")

                # Verificar si es Dia laborable (si es True, entonces NO es vÃ¡lido para guardias)
                es_laborable = es_dia_laborable(fecha_guardia)
                logger.info(
                    f"Â¿Es Dia laborable {fecha_guardia}? {es_laborable}")

                if es_laborable:
                    # Es Dia laborable (lunes a viernes normal), NO permitido para guardias
                    logger.warning(
                        f"Rechazando guardia en Dia laborable: {fecha_guardia}")
                    return Response(
                        {'error': 'Las guardias solo pueden programarse en fines de semana (sÃ¡bados y domingos) o feriados'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Si llegamos aquÃ­, es fin de semana o feriado (vÃ¡lido para guardias)
                motivo = get_motivo_no_laborable(fecha_guardia)
                logger.info(
                    f"âœ… Fecha vÃ¡lida para guardia {fecha_guardia}: {motivo}")

            except ValueError as e:
                return Response(
                    {'error': f'Formato de fecha invÃ¡lido: {data.get("fecha")}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            except Exception as e:
                logger.error(f"Error validando Dia laborable: {str(e)}")
                return Response(
                    {'error': 'Error validando fecha de guardia'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            # Obtener el agente del usuario autenticado
            from .utils import get_agente_rol, requiere_aprobacion_rol

            agente_id = data.get('agente_id')  # Por ahora recibir del request
            print(f"ðŸ” agente_id del request: {agente_id}")

            if not agente_id and hasattr(request.user, 'agente'):
                agente_id = request.user.agente.id_agente

            # Si no se proporciona agente_id, usar el de la sesiÃ³n
            if not agente_id:
                agente_id = request.session.get('user_id')
                print(f"ðŸ“Œ agente_id desde sesiÃ³n: {agente_id}")

            if not agente_id:
                return Response({'error': 'No se pudo determinar el agente creador'}, status=status.HTTP_400_BAD_REQUEST)

            # RBAC: ValidaciÃ³n de Ã¡rea permitida segun rol
            agente_creador = obtener_agente_sesion(request)
            if agente_creador:
                rol_creador = obtener_rol_agente(agente_creador)
                id_area_cronograma = data.get('id_area')

                if rol_creador == 'jefatura':
                    # Jefatura solo puede crear para su propia Ã¡rea
                    if agente_creador.id_area and agente_creador.id_area.id_area != id_area_cronograma:
                        return Response({
                            'error': f'Jefatura solo puede crear cronogramas para su propia Ã¡rea (Ãrea: {agente_creador.id_area.nombre})'
                        }, status=status.HTTP_403_FORBIDDEN)

                elif rol_creador == 'director':
                    # Director solo para Ã¡reas bajo su direcciÃ³n
                    areas_permitidas = obtener_areas_jerarquia(agente_creador)
                    area_ids = [a.id_area for a in areas_permitidas]
                    if id_area_cronograma not in area_ids:
                        return Response({
                            'error': 'Director solo puede crear cronogramas para Ã¡reas bajo su direcciÃ³n'
                        }, status=status.HTTP_403_FORBIDDEN)

                # Admin: sin restricciÃ³n

            # Obtener el agente
            print(f"âœ… Usando agente_id: {agente_id}")

            # Obtener agente completo
            try:
                from personas.models import Agente
                agente_creador = Agente.objects.get(id_agente=agente_id)
            except Agente.DoesNotExist:
                return Response(
                    {'error': 'Agente no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Determinar rol del creador
            rol_creador = get_agente_rol(agente_creador)
            if not rol_creador:
                rol_creador = 'jefatura'  # Por defecto

            # Determinar estado inicial segun rol
            if rol_creador.lower() == 'administrador':
                estado_inicial = 'publicada'  # Auto-aprobado y publicado
                fecha_aprobacion = timezone.now().date()
                aprobado_por_id = agente_id  # Se aprueba a sÃ­ mismo
            else:
                estado_inicial = 'pendiente'  # Requiere aprobaciÃ³n
                fecha_aprobacion = None
                aprobado_por_id = None

            # Crear el cronograma
            cronograma = Cronograma.objects.create(
                id_area_id=data['id_area'],
                id_jefe_id=agente_id,  # El jefe es quien crea
                id_director_id=agente_id,  # Por ahora, el mismo agente es jefe y director
                tipo=data['tipo'],
                hora_inicio=data['hora_inicio'],
                hora_fin=data['hora_fin'],
                estado=estado_inicial,
                fecha_creacion=timezone.now().date(),
                fecha_aprobacion=fecha_aprobacion,
                creado_por_rol=rol_creador,
                creado_por_id_id=agente_id,
                aprobado_por_id_id=aprobado_por_id
            )

            # Registrar auditorÃ­a del cronograma
            Auditoria.objects.create(
                pk_afectada=cronograma.id_cronograma,
                nombre_tabla='cronograma',
                creado_en=timezone.now(),
                valor_previo=None,
                valor_nuevo={
                    'nombre': data['nombre'],
                    'tipo': data['tipo'],
                    'id_area': data['id_area'],
                    'hora_inicio': data['hora_inicio'],
                    'hora_fin': data['hora_fin'],
                    'observaciones': data.get('observaciones', '')
                },
                accion='CREAR',
                id_agente_id=agente_id
            )

            # Determinar estado de las guardias segun estado del cronograma
            if estado_inicial == 'pendiente':
                estado_guardias = 'pendiente_aprobacion'
                guardias_activas = False  # No activar hasta aprobar
            else:
                # Para admin (publicada) y otros estados aprobados
                estado_guardias = 'planificada'
                guardias_activas = True

            # Validar fecha y duracion de guardia antes de crear
            from datetime import datetime
            from .utils import ValidadorHorarios

            fecha_guardia = datetime.strptime(data['fecha'], '%Y-%m-%d').date()

            # Validar que la fecha sea apta para guardias
            fecha_valida, mensaje_fecha = ValidadorHorarios.validar_fecha_guardia(
                fecha_guardia)
            if not fecha_valida:
                return Response(
                    {'error': f'Fecha no vÃ¡lida para guardia: {mensaje_fecha}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar duracion de la guardia
            from datetime import datetime as dt
            hora_inicio_obj = dt.strptime(data['hora_inicio'], '%H:%M').time()
            hora_fin_obj = dt.strptime(data['hora_fin'], '%H:%M').time()

            duracion_valida, mensaje_duracion = ValidadorHorarios.validar_duracion_guardia(
                hora_inicio_obj, hora_fin_obj
            )
            if not duracion_valida:
                return Response(
                    {'error': f'DuraciÃ³n no vÃ¡lida: {mensaje_duracion}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            print(
                f"âœ… Validaciones pasadas - {mensaje_fecha}, {mensaje_duracion}")

            # Debug: Mostrar datos antes de crear guardias
            print(f"ðŸ” Datos para crear guardias:")
            print(f"  - Cronograma creado: {cronograma.id_cronograma}")
            print(f"  - Estado guardias: {estado_guardias}")
            print(f"  - Guardias activas: {guardias_activas}")
            print(f"  - Agentes a procesar: {data['agentes']}")

            # Crear las guardias para cada agente
            guardias_creadas = []
            for agente_data in data['agentes']:
                # Manejar tanto el formato de entero como de objeto
                if isinstance(agente_data, dict):
                    agente_id_guardia = agente_data['id_agente']
                else:
                    agente_id_guardia = agente_data

                print(
                    f"ðŸ” Creando guardia para agente: {agente_id_guardia} (tipo: {type(agente_data)})")

                guardia = Guardia.objects.create(
                    id_cronograma=cronograma,
                    id_agente_id=agente_id_guardia,
                    fecha=data['fecha'],
                    hora_inicio=data['hora_inicio'],
                    hora_fin=data['hora_fin'],
                    tipo=data['tipo'],
                    estado=estado_guardias,
                    activa=guardias_activas,
                    observaciones=data.get('observaciones', '')
                )
                guardias_creadas.append(guardia)
                print(
                    f"âœ… Guardia creada: {guardia.id_guardia} para agente {agente_id_guardia}")

                # Registrar auditorÃ­a de cada guardia
                Auditoria.objects.create(
                    pk_afectada=guardia.id_guardia,
                    nombre_tabla='guardia',
                    creado_en=timezone.now(),
                    valor_previo=None,
                    valor_nuevo={
                        'id_cronograma': cronograma.id_cronograma,
                        'id_agente': agente_id_guardia,
                        'fecha': data['fecha'],
                        'hora_inicio': data['hora_inicio'],
                        'hora_fin': data['hora_fin'],
                        'tipo': data['tipo'],
                        'estado': 'planificada'
                    },
                    accion='CREAR',
                    id_agente_id=agente_id
                )

            print(
                f"âœ… Todas las guardias creadas exitosamente. Total: {len(guardias_creadas)}")

            return Response({
                'mensaje': 'Guardia creada exitosamente',
                'cronograma_id': cronograma.id_cronograma,
                'guardias_creadas': len(guardias_creadas)
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Error al crear guardia con cronograma: {str(e)}")
            return Response(
                {'error': f'Error al crear guardia: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['put', 'patch'])
    def actualizar_con_guardias(self, request, pk=None):
        """Actualiza un cronograma y sus guardias con registro completo en auditorÃ­a"""
        try:
            cronograma = self.get_object()
            data = request.data

            # Validar que el cronograma pueda ser editado
            if cronograma.estado not in ['pendiente', 'aprobada', 'generada']:
                return Response(
                    {'error': f'No se puede editar un cronograma en estado "{cronograma.estado}"'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar datos requeridos
            required_fields = ['nombre', 'tipo', 'id_area',
                               'fecha', 'hora_inicio', 'hora_fin', 'agentes']
            for field in required_fields:
                if field not in data:
                    return Response(
                        {'error': f'Campo requerido: {field}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Validar que haya agentes
            if not data['agentes'] or len(data['agentes']) == 0:
                return Response(
                    {'error': 'Debe seleccionar al menos un agente'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar que la fecha sea fin de semana o feriado
            try:
                from datetime import datetime as dt
                fecha_guardia = dt.strptime(data['fecha'], '%Y-%m-%d').date()

                # Verificar si es Dia laborable (si es True, entonces NO es vÃ¡lido para guardias)
                es_laborable = es_dia_laborable(fecha_guardia)
                logger.info(
                    f"Â¿Es Dia laborable {fecha_guardia}? {es_laborable}")

                if es_laborable:
                    # Es Dia laborable (lunes a viernes normal), NO permitido para guardias
                    logger.warning(
                        f"Rechazando actualizaciÃ³n de guardia en Dia laborable: {fecha_guardia}")
                    return Response(
                        {'error': 'Las guardias solo pueden programarse en fines de semana (sÃ¡bados y domingos) o feriados'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                motivo = get_motivo_no_laborable(fecha_guardia)
                logger.info(
                    f"âœ… Fecha vÃ¡lida para actualizar guardia {fecha_guardia}: {motivo}")

            except ValueError as e:
                return Response(
                    {'error': f'Formato de fecha invÃ¡lido: {data.get("fecha")}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            except Exception as e:
                logger.error(f"Error validando Dia laborable: {str(e)}")
                return Response(
                    {'error': 'Error validando fecha de guardia'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            # Obtener agente que realiza la actualizaciÃ³n
            agente_id = data.get('agente_id')
            if not agente_id and hasattr(request.user, 'agente'):
                agente_id = request.user.agente.id_agente

            if not agente_id:
                return Response(
                    {'error': 'No se pudo determinar el agente que actualiza'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                from personas.models import Agente
                agente_actualizador = Agente.objects.get(id_agente=agente_id)
            except Agente.DoesNotExist:
                return Response(
                    {'error': 'Agente no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Guardar valores previos para auditorÃ­a
            valor_previo_cronograma = {
                'tipo': cronograma.tipo,
                'hora_inicio': str(cronograma.hora_inicio),
                'hora_fin': str(cronograma.hora_fin),
                'estado': cronograma.estado,
                'id_area': cronograma.id_area_id
            }

            # Obtener guardias existentes para auditorÃ­a
            guardias_previas = list(cronograma.guardia_set.all().values(
                'id_guardia', 'id_agente_id', 'fecha', 'hora_inicio',
                'hora_fin', 'tipo', 'estado', 'activa'
            ))

            # Actualizar campos del cronograma
            cronograma.tipo = data['tipo']
            cronograma.hora_inicio = data['hora_inicio']
            cronograma.hora_fin = data['hora_fin']
            cronograma.id_area_id = data['id_area']

            # Si estaba aprobado/publicado y se modifica, volver a pendiente
            if cronograma.estado in ['aprobada', 'publicada']:
                from .utils import get_agente_rol
                rol_actualizador = get_agente_rol(agente_actualizador)

                # Solo admin puede editar sin cambiar estado
                if rol_actualizador and rol_actualizador.lower() != 'administrador':
                    cronograma.estado = 'pendiente'
                    cronograma.fecha_aprobacion = None
                    cronograma.aprobado_por_id = None

            cronograma.save()

            # Registrar auditorÃ­a del cronograma
            Auditoria.objects.create(
                pk_afectada=cronograma.id_cronograma,
                nombre_tabla='cronograma',
                creado_en=timezone.now(),
                valor_previo=valor_previo_cronograma,
                valor_nuevo={
                    'tipo': cronograma.tipo,
                    'hora_inicio': str(cronograma.hora_inicio),
                    'hora_fin': str(cronograma.hora_fin),
                    'estado': cronograma.estado,
                    'id_area': cronograma.id_area_id
                },
                accion='ACTUALIZAR',
                id_agente_id=agente_id
            )

            # Eliminar guardias existentes y registrar en auditorÃ­a
            for guardia_previa in guardias_previas:
                # Convertir fechas a string para serializaciÃ³n JSON
                guardia_previa_serializable = guardia_previa.copy()
                if 'fecha' in guardia_previa_serializable:
                    guardia_previa_serializable['fecha'] = str(
                        guardia_previa_serializable['fecha'])
                if 'hora_inicio' in guardia_previa_serializable:
                    guardia_previa_serializable['hora_inicio'] = str(
                        guardia_previa_serializable['hora_inicio'])
                if 'hora_fin' in guardia_previa_serializable:
                    guardia_previa_serializable['hora_fin'] = str(
                        guardia_previa_serializable['hora_fin'])

                Auditoria.objects.create(
                    pk_afectada=guardia_previa['id_guardia'],
                    nombre_tabla='guardia',
                    creado_en=timezone.now(),
                    valor_previo=guardia_previa_serializable,
                    valor_nuevo=None,
                    accion='ELIMINAR',
                    id_agente_id=agente_id
                )

            cronograma.guardia_set.all().delete()

            # Determinar estado de las nuevas guardias
            if cronograma.estado == 'pendiente':
                estado_guardias = 'pendiente_aprobacion'
                guardias_activas = False
            else:
                # Para estados aprobados y publicados
                estado_guardias = 'planificada'
                guardias_activas = True

            # Crear las nuevas guardias
            guardias_creadas = []
            for agente_data in data['agentes']:
                if isinstance(agente_data, dict):
                    agente_id_guardia = agente_data['id_agente']
                else:
                    agente_id_guardia = agente_data

                guardia = Guardia.objects.create(
                    id_cronograma=cronograma,
                    id_agente_id=agente_id_guardia,
                    fecha=data['fecha'],
                    hora_inicio=data['hora_inicio'],
                    hora_fin=data['hora_fin'],
                    tipo=data['tipo'],
                    estado=estado_guardias,
                    activa=guardias_activas,
                    observaciones=data.get('observaciones', '')
                )
                guardias_creadas.append(guardia)

                # Registrar auditorÃ­a de cada nueva guardia
                Auditoria.objects.create(
                    pk_afectada=guardia.id_guardia,
                    nombre_tabla='guardia',
                    creado_en=timezone.now(),
                    valor_previo=None,
                    valor_nuevo={
                        'id_cronograma': cronograma.id_cronograma,
                        'id_agente': agente_id_guardia,
                        'fecha': str(data['fecha']),
                        'hora_inicio': str(data['hora_inicio']),
                        'hora_fin': str(data['hora_fin']),
                        'tipo': data['tipo'],
                        'estado': estado_guardias,
                        'activa': guardias_activas
                    },
                    accion='CREAR',
                    id_agente_id=agente_id
                )

            # Registrar resumen en auditorÃ­a
            Auditoria.objects.create(
                pk_afectada=cronograma.id_cronograma,
                nombre_tabla='cronograma',
                creado_en=timezone.now(),
                valor_previo={'guardias_previas_count': len(guardias_previas)},
                valor_nuevo={'guardias_nuevas_count': len(
                    guardias_creadas), 'actualizacion_completa': True},
                accion='ACTUALIZACION_COMPLETA',
                id_agente_id=agente_id
            )

            return Response({
                'mensaje': 'Cronograma actualizado exitosamente',
                'cronograma_id': cronograma.id_cronograma,
                'guardias_eliminadas': len(guardias_previas),
                'guardias_creadas': len(guardias_creadas),
                'estado_cronograma': cronograma.estado
            }, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'data_received': request.data
            }
            logger.error(
                f"Error al actualizar cronograma con guardias: {error_details}")
            return Response(
                {'error': f'Error al actualizar: {str(e)}',
                 'details': error_details},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def planificar(self, request):
        """PlanificaciÃ³n automÃ¡tica de guardias"""
        serializer = PlanificacionCronogramaSerializer(data=request.data)
        if serializer.is_valid():
            try:
                resultado = PlanificadorCronograma.planificar_automatico(
                    area_id=serializer.validated_data['area_id'],
                    fecha_inicio=serializer.validated_data['fecha_inicio'],
                    fecha_fin=serializer.validated_data['fecha_fin'],
                    agentes_ids=serializer.validated_data.get('agentes_ids')
                )

                if resultado:
                    return Response(resultado, status=status.HTTP_201_CREATED)
                else:
                    return Response(
                        {'error': 'Error en la planificaciÃ³n automÃ¡tica'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )

            except Exception as e:
                logger.error(f"Error en planificaciÃ³n: {e}")
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['patch'])
    def aprobar(self, request, pk=None):
        """Aprueba un cronograma con validaciÃ³n de jerarquÃ­a de roles"""
        from .utils import get_agente_rol, puede_aprobar

        cronograma = self.get_object()

        # Validar estado actual
        if cronograma.estado not in ['generada', 'pendiente']:
            return Response(
                {'error': f'Solo se pueden aprobar cronogramas en estado "generada" o "pendiente". Estado actual: {cronograma.estado}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener agente que aprueba (simular con agente_id del request por ahora)
        # En producciÃ³n usar: agente_aprobador = request.user.agente
        agente_id = request.data.get('agente_id')
        if not agente_id:
            return Response(
                {'error': 'Se requiere agente_id para aprobar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from personas.models import Agente
            agente_aprobador = Agente.objects.get(id_agente=agente_id)
        except Agente.DoesNotExist:
            return Response(
                {'error': 'Agente no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Obtener rol del aprobador
        rol_aprobador = get_agente_rol(agente_aprobador)
        if not rol_aprobador:
            return Response(
                {'error': 'El agente no tiene un rol asignado'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validar si puede aprobar segun jerarquÃ­a
        if not puede_aprobar(cronograma, rol_aprobador):
            roles_permitidos = cronograma.puede_aprobar_rol
            return Response(
                {
                    'error': f'No tiene permisos para aprobar este cronograma',
                    'detalle': f'Roles permitidos: {", ".join(roles_permitidos)}. Su rol: {rol_aprobador}'
                },
                status=status.HTTP_403_FORBIDDEN
            )

        # Aprobar y publicar cronograma directamente
        cronograma.estado = 'publicada'
        cronograma.fecha_aprobacion = date.today()
        cronograma.aprobado_por_id = agente_aprobador
        cronograma.save()

        # Activar guardias asociadas que estaban pendientes
        guardias_activadas = cronograma.guardia_set.filter(
            estado='pendiente_aprobacion'
        ).update(
            estado='planificada',
            activa=True
        )

        # Registrar en auditorÃ­a
        Auditoria.objects.create(
            pk_afectada=cronograma.id_cronograma,
            nombre_tabla='cronograma',
            creado_en=timezone.now(),
            valor_previo={'estado': 'pendiente'},
            valor_nuevo={'estado': 'publicada',
                         'guardias_activadas': guardias_activadas},
            accion='APROBAR_Y_PUBLICAR',
            id_agente_id=agente_id
        )

        return Response({
            'mensaje': 'Cronograma aprobado y publicado exitosamente',
            'cronograma_id': cronograma.id_cronograma,
            'aprobado_por': f'{agente_aprobador.nombre} {agente_aprobador.apellido}',
            'fecha_aprobacion': cronograma.fecha_aprobacion,
            'guardias_activadas': guardias_activadas
        })

    @action(detail=True, methods=['patch'])
    def publicar(self, request, pk=None):
        """Publica un cronograma pendiente o aprobado (mÃ©todo legacy)"""
        cronograma = self.get_object()

        if cronograma.estado not in ['aprobada', 'pendiente']:
            return Response(
                {'error': f'Solo se pueden publicar cronogramas pendientes o aprobados. Estado actual: {cronograma.estado}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        estado_previo = cronograma.estado
        cronograma.estado = 'publicada'
        cronograma.save()

        # Registrar en auditorÃ­a
        Auditoria.objects.create(
            pk_afectada=cronograma.id_cronograma,
            nombre_tabla='cronograma',
            creado_en=timezone.now(),
            valor_previo={'estado': estado_previo},
            valor_nuevo={'estado': 'publicada'},
            accion='PUBLICAR_DIRECTO',
            id_agente_id=request.data.get('agente_id', 1)
        )

        return Response({'mensaje': 'Cronograma publicado exitosamente'})

    @action(detail=True, methods=['patch'])
    def despublicar(self, request, pk=None):
        """Despublica un cronograma para permitir ediciÃ³n"""
        cronograma = self.get_object()

        if cronograma.estado != 'publicada':
            return Response(
                {'error': 'Solo se pueden despublicar cronogramas publicados'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener agente que despublica
        agente_id = request.data.get('agente_id')
        if not agente_id:
            return Response(
                {'error': 'Se requiere agente_id para despublicar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from personas.models import Agente
            agente_despublicador = Agente.objects.get(id_agente=agente_id)
        except Agente.DoesNotExist:
            return Response(
                {'error': 'Agente no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Cambiar estado a pendiente (permite editar y eliminar)
        estado_previo = cronograma.estado
        cronograma.estado = 'pendiente'
        cronograma.save()

        # Registrar en auditorÃ­a
        Auditoria.objects.create(
            pk_afectada=cronograma.id_cronograma,
            nombre_tabla='cronograma',
            creado_en=timezone.now(),
            valor_previo={'estado': estado_previo},
            valor_nuevo={'estado': 'pendiente'},
            accion='DESPUBLICAR',
            id_agente_id=agente_id
        )

        return Response({
            'mensaje': 'Cronograma despublicado exitosamente. Ahora esta¡ pendiente y puede editarse o eliminarse.',
            'cronograma_id': cronograma.id_cronograma,
            'nuevo_estado': cronograma.estado
        })

    @action(detail=True, methods=['delete'])
    def eliminar(self, request, pk=None):
        """Elimina un cronograma solo si esta¡ en estado pendiente"""
        cronograma = self.get_object()

        if cronograma.estado != 'pendiente':
            return Response(
                {'error': 'Solo se pueden eliminar cronogramas en estado pendiente'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener agente que elimina
        agente_id = request.data.get(
            'agente_id') or request.query_params.get('agente_id')
        if not agente_id:
            return Response(
                {'error': 'Se requiere agente_id para eliminar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from personas.models import Agente
            agente_eliminador = Agente.objects.get(id_agente=agente_id)
        except Agente.DoesNotExist:
            return Response(
                {'error': 'Agente no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Guardar datos para auditorÃ­a antes de eliminar
        cronograma_id = cronograma.id_cronograma
        cronograma_data = {
            'id': cronograma_id,
            'nombre': cronograma.nombre,
            'estado': cronograma.estado,
            'tipo': cronograma.tipo,
            'fecha': str(cronograma.fecha),
        }

        # Registrar eliminaciÃ³n de guardias asociadas
        guardias = cronograma.guardia_set.all()
        for guardia in guardias:
            Auditoria.objects.create(
                pk_afectada=guardia.id_guardia,
                nombre_tabla='guardia',
                creado_en=timezone.now(),
                valor_previo={
                    'id_guardia': guardia.id_guardia,
                    'id_cronograma': cronograma_id,
                    'id_agente': guardia.id_agente_id
                },
                valor_nuevo=None,
                accion='ELIMINAR',
                id_agente_id=agente_id
            )

        # Registrar eliminaciÃ³n del cronograma
        Auditoria.objects.create(
            pk_afectada=cronograma_id,
            nombre_tabla='cronograma',
            creado_en=timezone.now(),
            valor_previo=cronograma_data,
            valor_nuevo=None,
            accion='ELIMINAR',
            id_agente_id=agente_id
        )

        # Eliminar cronograma (esto tambiÃ©n elimina las guardias por CASCADE)
        cronograma.delete()

        return Response({
            'mensaje': 'Cronograma eliminado exitosamente',
            'cronograma_eliminado': cronograma_data
        })

    @action(detail=False, methods=['get', 'post'])
    def pendientes(self, request):
        """Lista cronogramas pendientes de aprobaciÃ³n segun rol del usuario"""
        from .utils import get_agente_rol, get_approval_hierarchy

        # Obtener agente del usuario (por ahora via query param)
        agente_id = request.query_params.get('agente_id')
        if not agente_id:
            if hasattr(request.user, 'agente'):
                agente_id = request.user.agente.id_agente
            else:
                return Response(
                    {'error': 'No se pudo determinar el agente'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        try:
            from personas.models import Agente
            agente = Agente.objects.get(id_agente=agente_id)
        except Agente.DoesNotExist:
            return Response(
                {'error': 'Agente no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Obtener rol del agente
        rol_agente = get_agente_rol(agente)
        if not rol_agente:
            return Response(
                {'error': 'El agente no tiene un rol asignado'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Filtrar cronogramas pendientes que el agente puede aprobar
        cronogramas_pendientes = []
        queryset = Cronograma.objects.filter(estado='pendiente').select_related(
            'id_area', 'id_jefe', 'creado_por_id'
        ).prefetch_related('guardia_set')

        for cronograma in queryset:
            # Verificar si el rol del agente puede aprobar este cronograma
            if cronograma.creado_por_rol:
                roles_permitidos = get_approval_hierarchy(
                    cronograma.creado_por_rol)
                if rol_agente.lower() in roles_permitidos:
                    cronogramas_pendientes.append(cronograma)
            elif rol_agente.lower() == 'administrador':
                # Si no tiene creado_por_rol, solo admin puede aprobar
                cronogramas_pendientes.append(cronograma)

        # Serializar resultados
        serializer = CronogramaExtendidoSerializer(
            cronogramas_pendientes, many=True)

        return Response({
            'count': len(cronogramas_pendientes),
            'rol_agente': rol_agente,
            'cronogramas': serializer.data
        })

    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        """Rechaza un cronograma con motivo"""
        from .utils import get_agente_rol, puede_aprobar

        cronograma = self.get_object()

        # Validar estado actual
        if cronograma.estado not in ['generada', 'pendiente']:
            return Response(
                {'error': f'Solo se pueden rechazar cronogramas pendientes. Estado actual: {cronograma.estado}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener motivo de rechazo
        motivo = request.data.get('motivo')
        if not motivo:
            return Response(
                {'error': 'Se requiere un motivo de rechazo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener agente que rechaza
        agente_id = request.data.get('agente_id')
        if not agente_id:
            return Response(
                {'error': 'Se requiere agente_id'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from personas.models import Agente
            agente_rechazador = Agente.objects.get(id_agente=agente_id)
        except Agente.DoesNotExist:
            return Response(
                {'error': 'Agente no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Validar permisos
        rol_rechazador = get_agente_rol(agente_rechazador)
        if not rol_rechazador:
            return Response(
                {'error': 'El agente no tiene un rol asignado'},
                status=status.HTTP_403_FORBIDDEN
            )

        if not puede_aprobar(cronograma, rol_rechazador):
            return Response(
                {'error': 'No tiene permisos para rechazar este cronograma'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Guardar estado previo para auditorÃ­a
        estado_previo = cronograma.estado

        # Rechazar cronograma
        cronograma.estado = 'rechazada'
        cronograma.save()

        # Registrar en auditorÃ­a
        Auditoria.objects.create(
            pk_afectada=cronograma.id_cronograma,
            nombre_tabla='cronograma',
            creado_en=timezone.now(),
            valor_previo={'estado': estado_previo},
            valor_nuevo={'estado': 'rechazada', 'motivo': motivo},
            accion='RECHAZO',
            id_agente_id=agente_id
        )

        return Response({
            'mensaje': 'Cronograma rechazado',
            'cronograma_id': cronograma.id_cronograma,
            'rechazado_por': f'{agente_rechazador.nombre} {agente_rechazador.apellido}',
            'motivo': motivo
        })


class GuardiaViewSet(viewsets.ModelViewSet):
    """ViewSet para guardias con reportes y exportaciÃ³n"""

    queryset = Guardia.objects.all()
    serializer_class = GuardiaResumenSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        """Override create para validar Dias permitidos"""
        fecha_str = request.data.get('fecha')
        if fecha_str:
            try:
                from datetime import datetime as dt
                fecha_obj = dt.strptime(fecha_str, '%Y-%m-%d').date()

                # Validar que NO sea Dia laborable (solo permitir fines de semana y feriados)
                if es_dia_laborable(fecha_obj):
                    return Response(
                        {'error': 'Las guardias solo pueden programarse en fines de semana (sÃ¡bados y domingos) o feriados'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            except ValueError:
                return Response(
                    {'error': 'Formato de fecha invÃ¡lido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        """Override update para validar Dias permitidos"""
        fecha_str = request.data.get('fecha')
        if fecha_str:
            try:
                from datetime import datetime as dt
                fecha_obj = dt.strptime(fecha_str, '%Y-%m-%d').date()

                # Validar que NO sea Dia laborable (solo permitir fines de semana y feriados)
                if es_dia_laborable(fecha_obj):
                    return Response(
                        {'error': 'Las guardias solo pueden programarse en fines de semana (sÃ¡bados y domingos) o feriados'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            except ValueError:
                return Response(
                    {'error': 'Formato de fecha invÃ¡lido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        """Override partial_update para validar Dias permitidos"""
        fecha_str = request.data.get('fecha')
        if fecha_str:
            try:
                from datetime import datetime as dt
                fecha_obj = dt.strptime(fecha_str, '%Y-%m-%d').date()

                # Validar que NO sea Dia laborable (solo permitir fines de semana y feriados)
                if es_dia_laborable(fecha_obj):
                    return Response(
                        {'error': 'Las guardias solo pueden programarse en fines de semana (sÃ¡bados y domingos) o feriados'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            except ValueError:
                return Response(
                    {'error': 'Formato de fecha invÃ¡lido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return super().partial_update(request, *args, **kwargs)

    @action(detail=True, methods=['get', 'post'], url_path='notas')
    def notas_guardia(self, request, pk=None):
        """
        GET: Lista notas de una guardia
        POST: Crea o actualiza nota personal del agente para esta guardia
        """
        guardia = self.get_object()

        if request.method == 'GET':
            # Listar todas las notas de esta guardia
            notas = guardia.notas.all().order_by('-fecha_nota')
            from .serializers import NotaGuardiaSerializer
            serializer = NotaGuardiaSerializer(notas, many=True)
            return Response(serializer.data)

        elif request.method == 'POST':
            # Crear o actualizar nota del agente actual
            agente_id = request.data.get(
                'id_agente') or request.data.get('agente_id')
            if not agente_id:
                return Response(
                    {'error': 'Se requiere id_agente o agente_id'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            from .models import NotaGuardia
            from .serializers import NotaGuardiaSerializer

            # Buscar si ya existe una nota de este agente para esta guardia
            nota_existente = NotaGuardia.objects.filter(
                id_guardia=guardia,
                id_agente_id=agente_id
            ).first()

            if nota_existente:
                # Actualizar nota existente
                serializer = NotaGuardiaSerializer(
                    nota_existente,
                    data=request.data,
                    partial=True
                )
            else:
                # Crear nueva nota
                data = request.data.copy()
                data['id_guardia'] = guardia.id_guardia
                serializer = NotaGuardiaSerializer(data=data)

            if serializer.is_valid():
                serializer.save()
                return Response(
                    serializer.data,
                    status=status.HTTP_201_CREATED if not nota_existente else status.HTTP_200_OK
                )

            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['put', 'delete'], url_path='notas/(?P<nota_id>[^/.]+)')
    def nota_detalle(self, request, nota_id=None, pk=None):
        """
        PUT: Actualiza una nota existente
        DELETE: Elimina una nota
        """
        from .models import NotaGuardia
        from .serializers import NotaGuardiaSerializer

        try:
            nota = NotaGuardia.objects.get(id_nota=nota_id)
        except NotaGuardia.DoesNotExist:
            return Response(
                {'error': 'Nota no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Validar que el agente que modifica sea el dueno de la nota
        agente_id = request.data.get(
            'agente_id') if request.method == 'PUT' else request.query_params.get('agente_id')
        if agente_id and str(nota.id_agente_id) != str(agente_id):
            return Response(
                {'error': 'No tiene permisos para modificar esta nota'},
                status=status.HTTP_403_FORBIDDEN
            )

        if request.method == 'PUT':
            serializer = NotaGuardiaSerializer(
                nota, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            nota.delete()
            return Response(
                {'mensaje': 'Nota eliminada exitosamente'},
                status=status.HTTP_204_NO_CONTENT
            )

    @action(detail=False, methods=['get', 'post'])
    def verificar_disponibilidad(self, request):
        """Verifica disponibilidad de un agente en una fecha especÃ­fica"""
        agente_id = data_in.get('agente')
        fecha = request.query_params.get('fecha')

        if not agente_id or not fecha:
            return Response(
                {'error': 'Se requieren parÃ¡metros: agente y fecha'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from datetime import datetime
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()

            # Verificar si el agente ya tiene guardias ese Dia
            guardias_existentes = self.get_queryset().filter(
                id_agente=agente_id,
                fecha=fecha_obj,
                activa=True
            ).count()

            disponible = guardias_existentes == 0

            return Response({
                'disponible': disponible,
                'guardias_existentes': guardias_existentes,
                'fecha': fecha,
                'agente_id': int(agente_id)
            })

        except ValueError as e:
            return Response(
                {'error': f'Fecha invÃ¡lida: {fecha}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'Error verificando disponibilidad: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def resumen(self, request):
        """Resumen de guardias por perÃ­odo - solo guardias activas y aprobadas"""
        fecha_desde = data_in.get('fecha_desde')
        fecha_hasta = data_in.get('fecha_hasta')
        agente_id = data_in.get('agente')
        area_id = data_in.get('area')

        queryset = self.get_queryset()

        # FILTRO CRÃTICO: Solo mostrar guardias activas de cronogramas aprobados/publicados
        # Esto previene que usuarios vean guardias pendientes de aprobaciÃ³n
        queryset = queryset.filter(
            activa=True,
            estado='planificada',
            id_cronograma__estado__in=['aprobada', 'publicada']
        )

        if fecha_desde:
            queryset = queryset.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha__lte=fecha_hasta)
        if agente_id:
            queryset = queryset.filter(id_agente=agente_id)
        if area_id:
            queryset = queryset.filter(id_cronograma__id_area=area_id)

        # EstadÃ­sticas
        estadisticas = queryset.aggregate(
            total_guardias=Count('id_guardia'),
            guardias_activas=Count('id_guardia', filter=Q(activa=True)),
            horas_planificadas=Sum('horas_planificadas'),
            horas_efectivas=Sum('horas_efectivas')
        )

        # Guardias del perÃ­odo
        # Limitar a 50 mÃ¡s recientes
        guardias = queryset.order_by('-fecha')[:50]
        serializer = self.get_serializer(guardias, many=True)

        return Response({
            'estadisticas': estadisticas,
            'guardias': serializer.data,
            'filtros_aplicados': {
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'agente_id': agente_id,
                'area_id': area_id
            }
        })

    @action(detail=False, methods=['get', 'post'])
    def guardias_por_cronograma(self, request):
        """Obtiene todas las guardias de un cronograma especifico (incluyendo pendientes)"""
        cronograma_id = request.query_params.get('id_cronograma')

        if not cronograma_id:
            return Response(
                {'error': 'Se requiere el parÃ¡metro id_cronograma'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Obtener todas las guardias del cronograma, sin filtrar por estado del cronograma
            guardias = self.get_queryset().filter(
                id_cronograma=cronograma_id,
                activa=True
            ).select_related('id_agente', 'id_cronograma').order_by('fecha', 'hora_inicio')

            # Serializar guardias con informaciÃ³n del agente
            guardias_data = []
            for guardia in guardias:
                agente = guardia.id_agente
                guardias_data.append({
                    'id_guardia': guardia.id_guardia,
                    'fecha': guardia.fecha.strftime('%Y-%m-%d'),
                    'hora_inicio': guardia.hora_inicio.strftime('%H:%M:%S') if guardia.hora_inicio else '',
                    'hora_fin': guardia.hora_fin.strftime('%H:%M:%S') if guardia.hora_fin else '',
                    'estado': guardia.estado,
                    'horas_planificadas': guardia.horas_planificadas,
                    'horas_efectivas': guardia.horas_efectivas,
                    'observaciones': guardia.observaciones,
                    'agente_id': agente.id_agente if agente else None,
                    'agente_nombre': f"{agente.nombre} {agente.apellido}" if agente else 'Sin asignar',
                    'agente_legajo': agente.legajo if agente else ''
                })

            return Response(guardias_data)

        except Exception as e:
            logger.error(
                f"Error obteniendo guardias del cronograma {cronograma_id}: {e}")
            return Response(
                {'error': f'Error obteniendo guardias del cronograma: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def reporte_individual(self, request):
        """Genera reporte individual segun documentacion - Planilla Individual de Guardias"""
        data_in = request.data if request.method.lower() == "post" else request.query_params
        filtros = {
            'agente': data_in.get('agente'),
            'area': data_in.get('area'),
            'fecha_desde': data_in.get('fecha_desde'),
            'fecha_hasta': data_in.get('fecha_hasta'),
            'tipo_guardia': data_in.get('tipo_guardia')
        }

        agente_sesion = obtener_agente_sesion(request)
        if not agente_sesion:
            return Response({'error': 'SesiÃ³n invÃ¡lida'}, status=status.HTTP_401_UNAUTHORIZED)

        user_ctx = {
            'agente': agente_sesion,
            'rol': obtener_rol_agente(agente_sesion)
        }

        try:
            data = obtener_datos_reporte(filtros, 'individual', user_ctx)
            return Response(data)
        except ReporteError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error generando reporte individual: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def reporte_general(self, request):
        """Genera reporte general segun documentacion - Planilla General/Preventiva"""

        data_in = request.data if request.method.lower() == 'post' else request.query_params
        filtros = {
            'agente': data_in.get('agente'),
            'area': data_in.get('area'),
            'fecha_desde': data_in.get('fecha_desde'),
            'fecha_hasta': data_in.get('fecha_hasta'),
            'tipo_guardia': data_in.get('tipo_guardia'),
            'incluir_feriados': data_in.get('incluir_feriados'),
            'incluir_licencias': data_in.get('incluir_licencias')
        }

        agente_sesion = obtener_agente_sesion(request)
        if not agente_sesion:
            return Response({'error': 'Sesion invalida'}, status=status.HTTP_401_UNAUTHORIZED)

        user_ctx = {
            'agente': agente_sesion,
            'rol': obtener_rol_agente(agente_sesion)
        }

        try:
            data = obtener_datos_reporte(filtros, 'general', user_ctx)
            return Response(data)
        except ReporteError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error generando reporte general: {e}")
            return Response(
                {'error': f'Error generando reporte general: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def por_mes(self, request):
        """Obtiene guardias para un mes especifico (optimizado para calendario)"""
        anio = request.query_params.get('anio')
        mes = request.query_params.get('mes')

        if not anio or not mes:
            return Response(
                {'error': 'anio y mes son requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from datetime import date, timedelta
            # Fechas del mes
            primer_dia = date(int(anio), int(mes), 1)
            if int(mes) == 12:
                ultimo_dia = date(int(anio) + 1, 1, 1) - timedelta(days=1)
            else:
                ultimo_dia = date(int(anio), int(mes) + 1,
                                  1) - timedelta(days=1)

            # Guardias que inician en el mes
            guardias_mes = Guardia.objects.filter(
                fecha__gte=primer_dia,
                fecha__lte=ultimo_dia,
                activa=True
            ).select_related('id_agente', 'id_cronograma')

            fecha_anterior = primer_dia - timedelta(days=1)
            guardias_extension = Guardia.objects.filter(
                fecha=fecha_anterior,
                hora_inicio__gt=F('hora_fin'),
                activa=True
            ).select_related('id_agente', 'id_cronograma')

            # Combinar ambos querysets
            todas_guardias = guardias_mes.union(
                guardias_extension).order_by('fecha', 'hora_inicio')

            return Response({
                'anio': anio,
                'mes': mes,
                'primer_dia': primer_dia,
                'ultimo_dia': ultimo_dia,
                'total_guardias': todas_guardias.count(),
                'guardias': GuardiaResumenSerializer(todas_guardias, many=True).data
            })

        except ValueError as e:
            return Response(
                {'error': f'anio o mes invÃ¡lidos: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'Error obteniendo guardias: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    # ========================
    # ENDPOINTS DE EXPORTACION
    # ========================

    @action(detail=False, methods=["get", "post"])
    def exportar_pdf(self, request):
        """
        Genera y descarga un reporte en formato PDF con formato institucional
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A3, landscape
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer, PageBreak
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from django.http import HttpResponse
            from django.contrib.staticfiles import finders
            from datetime import datetime as dt
            from io import BytesIO

            from personas.models import Area

            # =========================
            # Helpers
            # =========================
            MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                     "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

            def _to_date(v):
                if v is None:
                    return None
                if hasattr(v, "year"):
                    return v
                return dt.strptime(v, "%Y-%m-%d").date()

            def draw_membrete(canvas, doc):
                canvas.saveState()

                page_w, page_h = doc.pagesize
                center_x = page_w / 2

                # === Logo ===
                logo_path = finders.find(
                    "logos/logoGobByN.jpeg") or finders.find("logos/logoGobColor.jpeg")
                logo_w = 35 * mm
                logo_h = 22 * mm

                # Pegado arriba
                top_pad = 6 * mm
                y_top = page_h - top_pad
                y_logo = y_top - logo_h

                if logo_path:
                    try:
                        x_logo = center_x - (logo_w / 2)
                        canvas.drawImage(
                            logo_path,
                            x_logo, y_logo,
                            width=logo_w, height=logo_h,
                            preserveAspectRatio=True,
                            mask="auto"
                        )
                    except Exception:
                        pass

                # Texto debajo del logo
                y_text = y_logo - 3 * mm
                canvas.setFont("Helvetica-Oblique", 9)
                canvas.drawCentredString(
                    center_x, y_text,
                    "Provincia de Tierra del Fuego, Antártida e Islas del Atlántico Sur"
                )

                y_text -= 4 * mm
                canvas.setFont("Helvetica", 9)
                canvas.drawCentredString(
                    center_x, y_text, "República Argentina")

                y_text -= 6 * mm
                canvas.setFont("Helvetica-Bold", 10)
                canvas.drawCentredString(
                    center_x, y_text, "SUBSECRETARÍA DE SEGURIDAD VIAL")

                # Línea separadora (por arriba del contenido)
                y_line = y_text - 4 * mm
                canvas.setLineWidth(0.6)
                canvas.line(doc.leftMargin, y_line,
                            page_w - doc.rightMargin, y_line)

                # Pie Malvinas (fijo abajo)
                footer_text = (
                    "Las Islas Malvinas, Georgias y Sándwich del Sur, "
                    "y los espacios marítimos e insulares correspondientes son Argentinos"
                )
                canvas.setFont("Helvetica-Oblique", 8)
                canvas.drawCentredString(page_w / 2, 10 * mm, footer_text)

                canvas.restoreState()

            # =========================
            # Request data
            # =========================
            tipo_reporte = request.data.get("tipo_reporte", "general")

            filtros = {
                "agente": request.data.get("agente"),
                "area": request.data.get("area"),
                "fecha_desde": request.data.get("fecha_desde"),
                "fecha_hasta": request.data.get("fecha_hasta"),
                "tipo_guardia": request.data.get("tipo_guardia"),
                "incluir_feriados": request.data.get("incluir_feriados"),
                "incluir_licencias": request.data.get("incluir_licencias"),
            }

            configuracion = request.data.get("configuracion", {})
            is_general = (tipo_reporte == "general")

            # =========================
            # Auth/context
            # =========================
            agente_sesion = obtener_agente_sesion(request)
            if not agente_sesion:
                return Response({"error": "Sesión inválida"}, status=status.HTTP_401_UNAUTHORIZED)

            user_ctx = {
                "agente": agente_sesion,
                "rol": obtener_rol_agente(agente_sesion),
            }

            datos_reporte = obtener_datos_reporte(
                filtros, tipo_reporte, user_ctx)

            # =========================
            # Doc setup
            # =========================
            buffer = BytesIO()
            page_size = landscape(A3) if is_general else A3

            doc = SimpleDocTemplate(
                buffer,
                pagesize=page_size,
                leftMargin=18 if is_general else 72,
                rightMargin=18 if is_general else 72,
                topMargin=55 * mm if is_general else 72,  # deja aire para membrete
                bottomMargin=18 if is_general else 72,
            )

            styles = getSampleStyleSheet()
            elements = []

            # =========================
            # Título institucional (1 sola vez)
            # =========================
            # Área (una sola vez)
            area_id = filtros.get("area")
            if isinstance(area_id, list):
                area_id = area_id[0] if area_id else None

            area_nombre = "Todas las Áreas"
            if area_id:
                try:
                    area_nombre = Area.objects.get(id_area=area_id).nombre
                except Area.DoesNotExist:
                    area_nombre = str(area_id)

            fd_title = _to_date(filtros.get("fecha_desde"))
            fh_title = _to_date(filtros.get("fecha_hasta"))

            # Texto meses (octubre–diciembre, 2025)
            mes_inicio = MESES[fd_title.month - 1] if fd_title else ""
            mes_fin = MESES[fh_title.month - 1] if fh_title else ""
            anio = fh_title.year if fh_title else ""

            if fd_title and fh_title and fd_title.month == fh_title.month and fd_title.year == fh_title.year:
                periodo_txt = f"mes {mes_inicio}, {anio}"
            else:
                periodo_txt = f"mes {mes_inicio}–{mes_fin}, {anio}"

            titulo_reporte = f"Planilla General {area_nombre} {periodo_txt}"

            title_style = ParagraphStyle(
                "Title",
                parent=styles["Heading1"],
                fontName="Helvetica-Bold",
                fontSize=14,
                alignment=1,
                spaceAfter=6,
            )

            elements.append(Paragraph(titulo_reporte, title_style))
            elements.append(Spacer(1, 6 * mm))

            # =========================
            # Tablas
            # =========================
            tabla_data = self._generar_tabla_pdf(
                tipo_reporte, datos_reporte, filtros)

            info_style = ParagraphStyle(
                "InfoStyle",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=9,
                spaceAfter=2,
            )

            if tipo_reporte == "general" and isinstance(tabla_data, dict) and "bloques" in tabla_data:
                bloques = tabla_data["bloques"]

                for i, bloque in enumerate(bloques):
                    anio = bloque["anio"]
                    mes = bloque["mes"]
                    rows = bloque["rows"]
                    weekend_cols = bloque["weekend_cols"]

                    # Período del mes (recortado al rango real)
                    import calendar
                    from datetime import date

                    last_day = calendar.monthrange(anio, mes)[1]
                    mes_inicio_dt = date(anio, mes, 1)
                    mes_fin_dt = date(anio, mes, last_day)

                    fd = _to_date(filtros.get("fecha_desde"))
                    fh = _to_date(filtros.get("fecha_hasta"))

                    periodo_inicio = max(fd, mes_inicio_dt)
                    periodo_fin = min(fh, mes_fin_dt)

                    # Info del mes (labels en negrita)
                    elements.append(
                        Paragraph(f"<b>Área:</b> {area_nombre}", info_style))
                    elements.append(Paragraph(
                        f"<b>Período:</b> {periodo_inicio.strftime('%d/%m/%Y')} al {periodo_fin.strftime('%d/%m/%Y')}",
                        info_style
                    ))
                    elements.append(Paragraph(
                        f"<b>Tipo de Guardia:</b> {filtros.get('tipo_guardia') or 'Regular'}",
                        info_style
                    ))
                    elements.append(Paragraph(
                        f"<b>Licencias:</b> {'Sí' if filtros.get('incluir_licencias') else 'No'} | "
                        f"<b>Feriados:</b> {'Sí' if filtros.get('incluir_feriados') else 'No'}",
                        info_style
                    ))
                    elements.append(Spacer(1, 6))

                    # Tabla
                    tabla = Table(rows)

                    base_style = [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),

                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]

                    weekend_fill = colors.HexColor("#7BE69A")
                    for col in weekend_cols:
                        base_style.append(
                            ("BACKGROUND", (col, 1), (col, -1), weekend_fill))

                    tabla.setStyle(TableStyle(base_style))
                    elements.append(tabla)

                    # Firmas solo al final del doc (no por mes)
                    if i < len(bloques) - 1:
                        elements.append(PageBreak())

            else:
                # Otros reportes (si los usás)
                if isinstance(tabla_data, tuple):
                    tabla_rows, weekend_cols = tabla_data
                else:
                    tabla_rows, weekend_cols = tabla_data, []

                tabla = Table(tabla_rows)
                tabla.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ]))
                elements.append(tabla)

            # =========================
            # Firmas (más abajo, cerca del pie)
            # =========================
            # ajustá: + = más abajo, - = más arriba
            elements.append(Spacer(1, 170))

            firma_data = [
                ["", ""],
                ["_" * 30, "_" * 30],
                ["Jefe de Área", "RR.HH./Liquidación"],
                ["Firma y Sello", "Firma y Sello"],
            ]

            firma_tabla = Table(firma_data, colWidths=[70 * mm, 70 * mm])
            firma_tabla.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
            ]))
            elements.append(firma_tabla)

            # =========================
            # Build PDF
            # =========================
            doc.build(elements, onFirstPage=draw_membrete,
                      onLaterPages=draw_membrete)

            buffer.seek(0)
            response = HttpResponse(buffer, content_type="application/pdf")

            timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
            filename = f"GIGA_{tipo_reporte}_{timestamp}.pdf"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except ReporteError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error generando PDF: {e}")
            return Response({"error": f"Error generando PDF: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _meses_en_rango(self, fd: date, fh: date):
        """Devuelve lista [(anio, mes), ...] desde fd hasta fh inclusive."""
        meses = []
        cur = fd.replace(day=1)
        end = fh.replace(day=1)
        while cur <= end:
            meses.append((cur.year, cur.month))
            if cur.month == 12:
                cur = cur.replace(year=cur.year + 1, month=1)
            else:
                cur = cur.replace(month=cur.month + 1)
        return meses

    def _generar_tabla_pdf(self, tipo_reporte, datos, filtros):
        """Genera datos de tabla especificos para cada tipo de reporte usando los datos reales."""
        if tipo_reporte == 'individual':
            headers = ['Fecha', 'Dia Semana', 'Horario Guardia',
                       'Horas Planificadas', 'Horas Efectivas', 'Motivo', 'Novedad']
            rows = [headers]

            for dia in datos.get('dias', []):
                horario_guardia = ""
                if dia.get('horario_guardia_inicio') and dia.get('horario_guardia_fin'):
                    horario_guardia = f"{dia['horario_guardia_inicio']}-{dia['horario_guardia_fin']}"

                rows.append([
                    dia.get('fecha', ''),
                    dia.get('dia_semana', ''),
                    horario_guardia,
                    dia.get('horas_planificadas', ''),
                    dia.get('horas_efectivas', ''),
                    dia.get('motivo_guardia', ''),
                    dia.get('novedad', ''),
                ])
            return rows

        # Reporte general: se arma como grilla Agente x Di­a
        if tipo_reporte == 'general':
            import calendar
            from datetime import datetime as dt

            fd = filtros.get("fecha_desde")
            fh = filtros.get("fecha_hasta")

            if not hasattr(fd, "year"):
                fd = dt.strptime(fd, "%Y-%m-%d").date()
            if not hasattr(fh, "year"):
                fh = dt.strptime(fh, "%Y-%m-%d").date()

            meses = self._meses_en_rango(fd, fh)

            bloques = []
            for anio, mes in meses:
                cant_dias = calendar.monthrange(anio, mes)[1]

                weekend_cols = []
                for day in range(1, cant_dias + 1):
                    dow = dt(anio, mes, day).weekday()
                    if dow in (5, 6):
                        weekend_cols.append(2 + (day - 1))

                headers = ['Apellido y Nombre', 'CUIL'] + \
                    [str(d) for d in range(1, 32)] + ['Total']
                rows = [headers]

                totales_por_dia = {d: 0 for d in range(1, 32)}
                total_general_mes = 0

                for agente in datos.get('agentes', []):
                    mapa = {}
                    for dia in agente.get("dias", []):
                        f = dia.get("fecha")
                        if not f:
                            continue
                        dtn = dt.strptime(f, "%Y-%m-%d")
                        if dtn.year == anio and dtn.month == mes:
                            mapa[dtn.day] = dia.get("valor", "")

                    fila = [agente.get('nombre_completo', ''),
                            agente.get('cuil', '')]
                    total_agente = 0

                    for d in range(1, 32):
                        if d <= cant_dias:
                            v = mapa.get(d, "")

                            if isinstance(v, (int, float)) and v > 0:
                                total_agente += v
                                totales_por_dia[d] += v
                                total_general_mes += v
                                fila.append(f"{int(v)} hs" if float(
                                    v).is_integer() else f"{v} hs")
                            else:
                                if isinstance(v, str) and v.strip():
                                    fila.append(v.strip())
                                else:
                                    fila.append("")
                        else:
                            fila.append("")

                    fila.append(f"{int(total_agente)} hs" if float(
                        total_agente).is_integer() else f"{total_agente} hs")
                    rows.append(fila)

                fila_total = ["Total", ""]
                for d in range(1, 32):
                    if d <= cant_dias:
                        td = totales_por_dia[d]
                        fila_total.append(f"{int(td)} hs" if td > 0 else "")
                    else:
                        fila_total.append("")
                fila_total.append(
                    f"{int(total_general_mes)} hs" if total_general_mes > 0 else "")
                rows.append(fila_total)

                bloques.append({
                    "anio": anio,
                    "mes": mes,
                    "rows": rows,
                    "weekend_cols": weekend_cols,
                    "cant_dias": cant_dias,
                })

            return {"bloques": bloques}

        # Fallback generico
        headers = ['Item', 'Descripcion', 'Valor']
        rows = [
            headers,
            ['Tipo de Reporte', tipo_reporte.replace('_', ' ').title(), ''],
            ['Periodo',
                f"{filtros.get('fecha_desde', '')} - {filtros.get('fecha_hasta', '')}", ''],
            ['Estado', 'Generado exitosamente', ''],
        ]
        return rows

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def exportar_csv(self, request):
        """
        Genera y descarga un reporte en formato CSV
        """
        try:
            import csv
            from django.http import HttpResponse
            from io import StringIO

            # Obtener datos de la request
            tipo_reporte = request.data.get('tipo_reporte', 'general')
            filtros = {
                'agente': request.data.get('agente'),
                'area': request.data.get('area'),
                'fecha_desde': request.data.get('fecha_desde'),
                'fecha_hasta': request.data.get('fecha_hasta'),
                'tipo_guardia': request.data.get('tipo_guardia')
            }
            configuracion = request.data.get('configuracion', {})

            agente_sesion = obtener_agente_sesion(request)
            if not agente_sesion:
                return Response({'error': 'SesiÃ³n invÃ¡lida'}, status=status.HTTP_401_UNAUTHORIZED)

            user_ctx = {
                'agente': agente_sesion,
                'rol': obtener_rol_agente(agente_sesion)
            }

            datos_reporte = obtener_datos_reporte(
                filtros, tipo_reporte, user_ctx)

            # Crear buffer para CSV
            output = StringIO()
            writer = csv.writer(output)

            # Escribir cabecera informativa
            writer.writerow(
                ['# Sistema GIGA - Reporte de Guardias y Asistencias'])
            writer.writerow(['# Universidad Nacional de Tierra del Fuego'])
            writer.writerow(
                [f'# Tipo de Reporte: {tipo_reporte.replace("_", " ").title()}'])
            writer.writerow(
                [f'# Perioodo: {filtros.get("fecha_desde", "")} - {filtros.get("fecha_hasta", "")}'])
            writer.writerow(
                [f'# Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
            writer.writerow([])  # Linea Vacia

            # Generar datos segun tipo de reporte
            datos_csv = self._generar_datos_csv(
                tipo_reporte, datos_reporte, filtros)

            # Escribir datos
            for fila in datos_csv:
                writer.writerow(fila)

            # Preparar respuesta
            response = HttpResponse(output.getvalue(), content_type='text/csv')

            # Generar nombre de archivo
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"GIGA_{tipo_reporte}_{timestamp}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except ReporteError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error generando CSV: {e}")
            return Response(
                {'error': f'Error generando CSV: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def exportar_excel(self, request):
        """
        Genera y descarga un reporte en formato Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            from io import BytesIO

            # Obtener datos de la request
            tipo_reporte = request.data.get('tipo_reporte', 'general')
            filtros = {
                'agente': request.data.get('agente'),
                'area': request.data.get('area'),
                'fecha_desde': request.data.get('fecha_desde'),
                'fecha_hasta': request.data.get('fecha_hasta'),
                'tipo_guardia': request.data.get('tipo_guardia')
            }

            agente_sesion = obtener_agente_sesion(request)
            if not agente_sesion:
                return Response({'error': 'SesiÃ³n invÃ¡lida'}, status=status.HTTP_401_UNAUTHORIZED)

            user_ctx = {
                'agente': agente_sesion,
                'rol': obtener_rol_agente(agente_sesion)
            }

            datos_reporte = obtener_datos_reporte(
                filtros, tipo_reporte, user_ctx)

            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Reporte {tipo_reporte.title()}"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(
                horizontal="center", vertical="center")

            # Cabecera informativa
            ws['A1'] = 'Sistema GIGA - Universidad Nacional de Tierra del Fuego'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Reporte: {tipo_reporte.replace("_", " ").title()}'
            ws['A3'] = f'Periodo: {filtros.get("fecha_desde", "")} - {filtros.get("fecha_hasta", "")}'
            ws['A4'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'

            # Linea Vacia
            row_start = 6

            # Generar datos
            datos_excel = self._generar_datos_csv(
                tipo_reporte, datos_reporte, filtros)

            # Escribir encabezados
            if datos_excel:
                headers = datos_excel[0]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_start, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment

                # Escribir datos
                for row_idx, fila in enumerate(datos_excel[1:], row_start + 1):
                    for col_idx, valor in enumerate(fila, 1):
                        ws.cell(row=row_idx, column=col_idx, value=valor)

                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Preparar respuesta
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # Generar nombre de archivo
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"GIGA_{tipo_reporte}_{timestamp}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except ReporteError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except ImportError:
            logger.error(
                "openpyxl no esta¡ instalado. Instale con: pip install openpyxl")
            return Response(
                {'error': 'Funcionalidad de Excel no disponible. Contacte al administrador.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            return Response(
                {'error': f'Error generando Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _generar_datos_csv(self, tipo_reporte, datos, filtros):
        """Genera datos en formato de filas para CSV/Excel usando los datos reales."""
        if tipo_reporte == 'individual':
            headers = ['Fecha', 'Dia Semana', 'Horario Guardia', 'Horas Planificadas',
                       'Horas Efectivas', 'Motivo', 'Novedad', 'Estado Asistencia']
            rows = [headers]

            for dia in datos.get('dias', []):
                horario_guardia = ""
                if dia.get('horario_guardia_inicio') and dia.get('horario_guardia_fin'):
                    horario_guardia = f"{dia['horario_guardia_inicio']}-{dia['horario_guardia_fin']}"

                rows.append([
                    dia.get('fecha', ''),
                    dia.get('dia_semana', ''),
                    horario_guardia,
                    dia.get('horas_planificadas', ''),
                    dia.get('horas_efectivas', ''),
                    dia.get('motivo_guardia', ''),
                    dia.get('novedad', ''),
                    dia.get('estado_asistencia', ''),
                ])
            return rows

        if tipo_reporte == 'general':
            dias = datos.get('dias_columnas', [])
            headers = ['Agente', 'Legajo'] + \
                [d.get('fecha', '') for d in dias] + ['Total Horas']
            rows = [headers]

            for agente in datos.get('agentes', []):
                fila = [
                    agente.get('nombre_completo', ''),
                    agente.get('legajo', '')
                ]
                for d in dias:
                    fecha = d.get('fecha')
                    celda = next((dia for dia in agente.get(
                        'dias', []) if dia.get('fecha') == fecha), {})
                    fila.append(celda.get('valor', ''))
                fila.append(agente.get('total_horas', ''))
                rows.append(fila)
            return rows

        # Formato generico
        headers = ['Descripcion', 'Valor', 'Observaciones']
        rows = [
            headers,
            ['Tipo de Reporte', tipo_reporte.replace(
                '_', ' ').title(), 'Generado automaticamente'],
            ['Periodo Consultado',
                f"{filtros.get('fecha_desde', '')} - {filtros.get('fecha_hasta', '')}", 'Rango de fechas seleccionado'],
            ['Fecha de Generacion', timezone.now().strftime("%d/%m/%Y %H:%M"),
             'Momento de creacion del reporte'],
            ['Sistema', 'GIGA - UNTDF', 'Universidad Nacional de Tierra del Fuego'],
        ]
        return rows


class ResumenGuardiaMesViewSet(viewsets.ModelViewSet):
    """ViewSet para resumen mensual con cálculo automático de plus"""

    queryset = ResumenGuardiaMes.objects.all()
    serializer_class = ResumenGuardiaMesExtendidoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtros
        mes = self.request.query_params.get('mes')
        anio = self.request.query_params.get('anio')
        agente_id = self.data_in.get('agente')
        estado_plus = self.request.query_params.get('estado_plus')

        if mes:
            queryset = queryset.filter(mes=mes)
        if anio:
            queryset = queryset.filter(anio=anio)
        if agente_id:
            queryset = queryset.filter(id_agente=agente_id)
        if estado_plus:
            queryset = queryset.filter(estado_plus=estado_plus)

        return queryset.order_by('-anio', '-mes')

    @action(detail=False, methods=['post'])
    def calcular_mensual(self, request):
        """Dispara el cálculo automático de plus mensual"""
        serializer = CalculoPlusSerializer(data=request.data)
        if serializer.is_valid():
            try:
                resultado = CalculadoraPlus.generar_asignaciones_plus(
                    mes=serializer.validated_data['mes'],
                    anio=serializer.validated_data['anio']
                )

                return Response(resultado, status=status.HTTP_200_OK)

            except Exception as e:
                logger.error(f"Error en cálculo mensual: {e}")
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['patch'])
    def aprobar_lote(self, request):
        """Aprueba un lote de asignaciones de plus"""
        serializer = AprobacionPlusSerializer(data=request.data)
        if serializer.is_valid():
            try:
                resumen_ids = serializer.validated_data['resumen_ids']

                # Actualizar los resúmenes
                resumenes = ResumenGuardiaMes.objects.filter(
                    id_resumen_guardia_mes__in=resumen_ids,
                    estado_plus='pendiente'
                )

                resumenes.update(
                    estado_plus='aprobado',
                    aprobado_en=timezone.now()
                )

                return Response({
                    'mensaje': f'{resumenes.count()} asignaciones de plus aprobadas',
                    'aprobados': resumen_ids
                })

            except Exception as e:
                logger.error(f"Error aprobando plus: {e}")
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get', 'post'])
    def reporte_plus_simplificado(self, request):
        """
        Genera reporte de plus usando las reglas simplificadas:
        - Área operativa + guardia = 40%
        - Otras áreas + 32+ horas = 40%
        - Resto = 20%
        """
        mes = request.query_params.get('mes')
        anio = request.query_params.get('anio')
        area_id = request.query_params.get('area_id')

        if not all([mes, anio]):
            return Response(
                {'error': 'Se requieren parámetros: mes, anio'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from personas.models import Agente, Area
            from .utils import CalculadoraPlus
            from django.db.models import Sum

            # Filtrar agentes
            agentes_query = Agente.objects.filter(activo=True)
            if area_id:
                agentes_query = agentes_query.filter(id_area_id=area_id)
                area = Area.objects.get(id_area=area_id)
                area_nombre = area.nombre
            else:
                area_nombre = "Todas las áreas"

            agentes_plus = []
            total_agentes_plus20 = 0
            total_agentes_plus40 = 0

            for agente in agentes_query:
                # Calcular plus usando nueva lógica
                porcentaje_plus = CalculadoraPlus.calcular_plus_simplificado(
                    agente.id_agente, int(mes), int(anio)
                )

                if porcentaje_plus > 0:
                    # Obtener detalles de guardias para el reporte
                    from .models import Guardia
                    from datetime import date

                    fecha_inicio = date(int(anio), int(mes), 1)
                    if int(mes) == 12:
                        fecha_fin = date(int(anio) + 1, 1, 1)
                    else:
                        fecha_fin = date(int(anio), int(mes) + 1, 1)

                    guardias = Guardia.objects.filter(
                        id_agente=agente.id_agente,
                        fecha__gte=fecha_inicio,
                        fecha__lt=fecha_fin,
                        activa=True,
                        estado='planificada'
                    )

                    total_horas = guardias.aggregate(
                        total=Sum('horas_efectivas')
                    )['total'] or 0

                    area_nombre_agente = agente.id_area.nombre if agente.id_area else "Sin área"
                    es_operativa = any(op in area_nombre_agente.lower() for op in [
                        'secretaría de protección civil', 'operativo', 'emergencias'
                    ])

                    # Determinar motivo del plus
                    if es_operativa and guardias.exists():
                        motivo = "Área operativa con guardias"
                    elif not es_operativa and total_horas >= 32:
                        motivo = f"Otras áreas con {total_horas}h (≥32h)"
                    else:
                        motivo = "Guardias con menos de 32h"

                    agentes_plus.append({
                        'agente_id': agente.id_agente,
                        'nombre_completo': f"{agente.apellido}, {agente.nombre}",
                        'legajo': agente.legajo,
                        'area_nombre': area_nombre_agente,
                        'es_area_operativa': es_operativa,
                        'total_horas_guardia': float(total_horas),
                        'cantidad_guardias': guardias.count(),
                        'porcentaje_plus': float(porcentaje_plus),
                        'motivo_plus': motivo
                    })

                    if porcentaje_plus >= 40:
                        total_agentes_plus40 += 1
                    else:
                        total_agentes_plus20 += 1

            resultado = {
                'area_nombre': area_nombre,
                'periodo': {
                    'mes': int(mes),
                    'anio': int(anio),
                    'mes_nombre': date(int(anio), int(mes), 1).strftime('%B %Y')
                },
                'agentes': agentes_plus,
                'resumen': {
                    'total_agentes_con_plus': len(agentes_plus),
                    'agentes_plus_20': total_agentes_plus20,
                    'agentes_plus_40': total_agentes_plus40,
                    'total_horas_todas_guardias': sum(a['total_horas_guardia'] for a in agentes_plus)
                },
                'reglas_aplicadas': {
                    'regla_1': "Área operativa + guardia = 40% plus",
                    'regla_2': "Otras áreas + 32+ horas = 40% plus",
                    'regla_3': "Resto con guardias = 20% plus"
                }
            }

            return Response({
                'success': True,
                'data': resultado
            })

        except Exception as e:
            logger.error(f"Error generando reporte plus: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class HoraCompensacionViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de horas de compensación por emergencias"""

    queryset = HoraCompensacion.objects.all()
    serializer_class = HoraCompensacionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Optimizar consultas
        queryset = queryset.select_related(
            'id_agente', 'id_guardia', 'id_cronograma',
            'solicitado_por', 'aprobado_por'
        )

        # Filtros
        agente_id = self.data_in.get('agente')
        estado = self.request.query_params.get('estado')
        mes = self.request.query_params.get('mes')
        anio = self.request.query_params.get('anio')
        pendientes = self.request.query_params.get('pendientes')

        if agente_id:
            queryset = queryset.filter(id_agente=agente_id)
        if estado:
            queryset = queryset.filter(estado=estado)
        if mes and anio:
            queryset = queryset.filter(
                fecha_servicio__month=mes,
                fecha_servicio__year=anio
            )
        if pendientes and pendientes.lower() == 'true':
            queryset = queryset.filter(estado='pendiente')

        return queryset.order_by('-fecha_servicio', '-creado_en')

    @action(detail=False, methods=['post'])
    def crear_compensacion(self, request):
        """Crea una nueva solicitud de compensación"""
        serializer = CrearCompensacionSerializer(data=request.data)
        if serializer.is_valid():
            try:
                from .utils import ValidadorHorarios
                from personas.models import Agente

                # Obtener datos validados
                data = serializer.validated_data
                guardia = data['guardia']
                agente = Agente.objects.get(id_agente=data['id_agente'])

                # Obtener agente que solicita (puede ser diferente al que trabajó)
                agente_solicitante_id = request.data.get(
                    'solicitado_por', data['id_agente'])
                agente_solicitante = Agente.objects.get(
                    id_agente=agente_solicitante_id)

                # Calcular valor monetario
                valor_hora, monto_total = ValidadorHorarios.calcular_valor_hora_compensacion(
                    agente, data['horas_extra']
                )

                # Crear la compensación
                compensacion = HoraCompensacion.objects.create(
                    id_agente=agente,
                    id_guardia=guardia,
                    id_cronograma=guardia.id_cronograma,
                    fecha_servicio=data['fecha_servicio'],
                    hora_inicio_programada=guardia.hora_inicio,
                    hora_fin_programada=guardia.hora_fin,
                    hora_fin_real=data['hora_fin_real'],
                    motivo=data['motivo'],
                    descripcion_motivo=data['descripcion_motivo'],
                    numero_acta=data.get('numero_acta', ''),
                    tipo_compensacion=data.get('tipo_compensacion', 'plus'),
                    solicitado_por=agente_solicitante,
                    valor_hora_extra=valor_hora,
                    monto_total=monto_total
                )

                # Registrar en auditoría
                Auditoria.objects.create(
                    pk_afectada=compensacion.id_hora_compensacion,
                    nombre_tabla='hora_compensacion',
                    creado_en=timezone.now(),
                    valor_previo=None,
                    valor_nuevo={
                        'agente': agente.nombre + ' ' + agente.apellido,
                        'fecha_servicio': str(data['fecha_servicio']),
                        'horas_extra': float(compensacion.horas_extra),
                        'motivo': data['motivo'],
                        'monto_total': float(monto_total)
                    },
                    accion='CREAR_COMPENSACION',
                    id_agente_id=agente_solicitante.id_agente
                )

                serializer_response = HoraCompensacionSerializer(compensacion)
                return Response({
                    'mensaje': 'Compensación creada exitosamente',
                    'compensacion': serializer_response.data
                }, status=status.HTTP_201_CREATED)

            except Exception as e:
                logger.error(f"Error creando compensación: {e}")
                return Response(
                    {'error': f'Error creando compensación: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def aprobar_lote(self, request):
        """Aprueba o rechaza un lote de compensaciones"""
        serializer = AprobacionCompensacionSerializer(data=request.data)
        if serializer.is_valid():
            try:
                from personas.models import Agente

                compensacion_ids = serializer.validated_data['compensacion_ids']
                accion = serializer.validated_data['accion']
                observaciones = serializer.validated_data.get(
                    'observaciones', '')

                # Obtener agente aprobador
                agente_id = request.data.get('agente_id')
                if not agente_id:
                    return Response(
                        {'error': 'Se requiere agente_id para aprobar'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                agente_aprobador = Agente.objects.get(id_agente=agente_id)

                # Obtener compensaciones
                compensaciones = HoraCompensacion.objects.filter(
                    id_hora_compensacion__in=compensacion_ids,
                    estado='pendiente'
                )

                procesadas = 0
                for compensacion in compensaciones:
                    try:
                        if accion == 'aprobar':
                            compensacion.aprobar(
                                agente_aprobador, observaciones)
                        else:
                            compensacion.rechazar(
                                agente_aprobador, observaciones)

                        # Registrar en auditoría
                        Auditoria.objects.create(
                            pk_afectada=compensacion.id_hora_compensacion,
                            nombre_tabla='hora_compensacion',
                            creado_en=timezone.now(),
                            valor_previo={'estado': 'pendiente'},
                            valor_nuevo={'estado': compensacion.estado},
                            accion='APROBAR_COMPENSACION' if accion == 'aprobar' else 'RECHAZAR_COMPENSACION',
                            id_agente_id=agente_aprobador.id_agente
                        )

                        procesadas += 1

                    except Exception as e:
                        logger.error(
                            f"Error procesando compensación {compensacion.id_hora_compensacion}: {e}")
                        continue

                return Response({
                    'mensaje': f'{procesadas} compensaciones procesadas exitosamente',
                    'accion': accion,
                    'procesadas': procesadas,
                    'total_solicitadas': len(compensacion_ids)
                })

            except Exception as e:
                logger.error(f"Error procesando compensaciones: {e}")
                return Response(
                    {'error': f'Error procesando compensaciones: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get', 'post'])
    def resumen_mensual(self, request):
        """Resumen mensual de compensaciones por agente"""
        serializer = ResumenCompensacionSerializer(
            data=request.query_params.dict())
        if serializer.is_valid():
            try:
                from personas.models import Agente

                agente_id = serializer.validated_data['agente']
                mes = serializer.validated_data['mes']
                anio = serializer.validated_data['anio']

                agente = Agente.objects.get(id_agente=agente_id)

                # Obtener resumen usando el método del modelo
                resumen = HoraCompensacion.resumen_mensual_agente(
                    agente, mes, anio)

                # Obtener compensaciones del mes
                compensaciones = self.get_queryset().filter(
                    id_agente=agente,
                    fecha_servicio__month=mes,
                    fecha_servicio__year=anio
                ).order_by('-fecha_servicio')

                compensaciones_data = HoraCompensacionSerializer(
                    compensaciones, many=True).data

                return Response({
                    'agente': {
                        'id': agente.id_agente,
                        'nombre_completo': f"{agente.apellido}, {agente.nombre}",
                        'legajo': agente.legajo
                    },
                    'periodo': {
                        'mes': mes,
                        'anio': anio,
                        'mes_nombre': date(anio, mes, 1).strftime('%B %Y')
                    },
                    'resumen': resumen,
                    'compensaciones': compensaciones_data
                })

            except Exception as e:
                logger.error(f"Error generando resumen mensual: {e}")
                return Response(
                    {'error': f'Error generando resumen: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get', 'post'])
    def reporte_compensaciones(self, request):
        """Reporte general de compensaciones por período y área"""
        mes = request.query_params.get('mes')
        anio = request.query_params.get('anio')
        area_id = request.query_params.get('area_id')
        estado = request.query_params.get('estado', 'aprobada')

        if not all([mes, anio]):
            return Response(
                {'error': 'Se requieren parámetros: mes, anio'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from personas.models import Area
            from decimal import Decimal

            queryset = self.get_queryset().filter(
                fecha_servicio__month=mes,
                fecha_servicio__year=anio,
                estado=estado
            )

            if area_id:
                queryset = queryset.filter(id_agente__id_area=area_id)
                area = Area.objects.get(id_area=area_id)
                area_nombre = area.nombre
            else:
                area_nombre = "Todas las áreas"

            # Agrupar por agente
            compensaciones_por_agente = {}
            total_horas_extra = Decimal('0')
            total_monto = Decimal('0')

            for compensacion in queryset:
                agente_id = compensacion.id_agente.id_agente
                if agente_id not in compensaciones_por_agente:
                    compensaciones_por_agente[agente_id] = {
                        'agente': {
                            'id': agente_id,
                            'nombre_completo': f"{compensacion.id_agente.apellido}, {compensacion.id_agente.nombre}",
                            'legajo': compensacion.id_agente.legajo,
                            'area': compensacion.id_agente.id_area.nombre if compensacion.id_agente.id_area else "Sin área"
                        },
                        'compensaciones': [],
                        'total_horas_extra': Decimal('0'),
                        'total_monto': Decimal('0')
                    }

                compensaciones_por_agente[agente_id]['compensaciones'].append({
                    'fecha_servicio': compensacion.fecha_servicio,
                    'horas_extra': compensacion.horas_extra,
                    'motivo': compensacion.get_motivo_display(),
                    'monto': compensacion.monto_total or Decimal('0')
                })

                compensaciones_por_agente[agente_id]['total_horas_extra'] += compensacion.horas_extra
                compensaciones_por_agente[agente_id]['total_monto'] += compensacion.monto_total or Decimal(
                    '0')

                total_horas_extra += compensacion.horas_extra
                total_monto += compensacion.monto_total or Decimal('0')

            # Estadísticas por motivo
            from django.db.models import Count, Sum
            stats_por_motivo = queryset.values('motivo').annotate(
                cantidad=Count('id_hora_compensacion'),
                total_horas=Sum('horas_extra')
            )

            resultado = {
                'area_nombre': area_nombre,
                'periodo': {
                    'mes': int(mes),
                    'anio': int(anio),
                    'mes_nombre': date(int(anio), int(mes), 1).strftime('%B %Y')
                },
                'estado_filtrado': estado,
                'agentes': list(compensaciones_por_agente.values()),
                'totales': {
                    'total_compensaciones': queryset.count(),
                    'total_agentes': len(compensaciones_por_agente),
                    'total_horas_extra': float(total_horas_extra),
                    'total_monto': float(total_monto),
                    'promedio_horas_agente': float(total_horas_extra / max(len(compensaciones_por_agente), 1))
                },
                'estadisticas_motivo': [
                    {
                        'motivo': item['motivo'],
                        'motivo_display': dict(HoraCompensacion.MOTIVO_CHOICES).get(item['motivo'], item['motivo']),
                        'cantidad': item['cantidad'],
                        'total_horas': float(item['total_horas'] or 0)
                    }
                    for item in stats_por_motivo
                ]
            }

            return Response({
                'success': True,
                'data': resultado
            })

        except Exception as e:
            logger.error(f"Error generando reporte compensaciones: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def crear_desde_guardia(self, request, pk=None):
        """Crea compensación directamente desde una guardia específica"""
        try:
            guardia = Guardia.objects.get(id_guardia=pk)

            # Validar que no exista ya una compensación para esta guardia
            compensacion_existente = HoraCompensacion.objects.filter(
                id_guardia=guardia
            ).exists()

            if compensacion_existente:
                return Response(
                    {'error': 'Ya existe una compensación para esta guardia'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar datos requeridos
            required_fields = ['hora_fin_real', 'motivo', 'descripcion_motivo']
            for field in required_fields:
                if field not in request.data:
                    return Response(
                        {'error': f'Campo requerido: {field}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Obtener agente solicitante
            agente_solicitante_id = request.data.get('solicitado_por')
            if not agente_solicitante_id:
                return Response(
                    {'error': 'Se requiere agente solicitante'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            from personas.models import Agente
            from datetime import datetime
            agente_solicitante = Agente.objects.get(
                id_agente=agente_solicitante_id)

            # Convertir hora_fin_real de string a objeto time
            hora_fin_str = request.data['hora_fin_real']
            if isinstance(hora_fin_str, str):
                hora_fin_real = datetime.strptime(hora_fin_str, '%H:%M').time()
            else:
                hora_fin_real = hora_fin_str

            # Crear compensación usando el método del modelo
            compensacion = HoraCompensacion.crear_desde_guardia_extendida(
                guardia=guardia,
                hora_fin_real=hora_fin_real,
                motivo=request.data['motivo'],
                descripcion=request.data['descripcion_motivo'],
                solicitado_por=agente_solicitante
            )

            serializer = HoraCompensacionSerializer(compensacion)
            return Response({
                'mensaje': 'Compensación creada desde guardia exitosamente',
                'compensacion': serializer.data
            }, status=status.HTTP_201_CREATED)

        except Guardia.DoesNotExist:
            return Response(
                {'error': 'Guardia no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error creando compensación desde guardia: {e}")
            return Response(
                {'error': f'Error creando compensación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def reporte_horas_trabajadas(self, request):
        """Reporte de Guardias y Compensaciones - Horas programadas vs efectivas"""
        area_id = data_in.get('area')
        fecha_desde = data_in.get('fecha_desde')
        fecha_hasta = data_in.get('fecha_hasta')

        if not all([fecha_desde, fecha_hasta]):
            return Response(
                {'error': 'Se requieren parámetros: fecha_desde, fecha_hasta'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from personas.models import Area
            from datetime import datetime

            fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()

            # Filtro base de guardias
            guardias_filter = {
                'fecha__range': [fecha_inicio, fecha_fin],
                'activa': True,
                'id_cronograma__estado__in': ['aprobada', 'publicada']
            }

            # Filtrar por área si se especifica
            if area_id:
                guardias_filter['id_cronograma__id_area'] = area_id
                area = Area.objects.get(id_area=area_id)
                area_nombre = area.nombre
            else:
                area_nombre = 'Todas las áreas'

            # Obtener guardias en el período
            guardias = self.get_queryset().filter(**guardias_filter)

            # Agrupar por agente
            agentes_data = {}
            for guardia in guardias:
                agente_id = guardia.id_agente.id_agente
                if agente_id not in agentes_data:
                    agentes_data[agente_id] = {
                        'agente': guardia.id_agente.nombre + ' ' + guardia.id_agente.apellido,
                        'legajo': guardia.id_agente.legajo,
                        'horas_programadas': 0,
                        'horas_efectivas': 0,
                        'guardias_fines_feriados': 0,
                        'total_guardias': 0
                    }

                # Calcular horas programadas
                if guardia.hora_inicio and guardia.hora_fin:
                    inicio = datetime.combine(
                        guardia.fecha, guardia.hora_inicio)
                    fin = datetime.combine(guardia.fecha, guardia.hora_fin)
                    if fin < inicio:
                        from datetime import timedelta
                        fin += timedelta(days=1)
                    horas = (fin - inicio).total_seconds() / 3600
                    agentes_data[agente_id]['horas_programadas'] += horas

                # Agregar horas efectivas si las tiene
                if guardia.horas_efectivas:
                    agentes_data[agente_id]['horas_efectivas'] += guardia.horas_efectivas

                # Contar guardias de fines de semana/feriados
                if guardia.fecha.weekday() >= 5:  # Sábado o domingo
                    agentes_data[agente_id]['guardias_fines_feriados'] += 1

                agentes_data[agente_id]['total_guardias'] += 1

            # Formatear respuesta
            agentes_reporte = []
            for agente_data in agentes_data.values():
                agente_data['horas_programadas'] = round(
                    agente_data['horas_programadas'], 2)
                agente_data['horas_efectivas'] = round(
                    agente_data['horas_efectivas'], 2)
                agente_data['total_horas'] = max(
                    agente_data['horas_programadas'], agente_data['horas_efectivas'])
                agentes_reporte.append(agente_data)

            # Ordenar por nombre
            agentes_reporte.sort(key=lambda x: x['agente'])

            return Response({
                'area_nombre': area_nombre,
                'periodo': {
                    'fecha_desde': fecha_desde,
                    'fecha_hasta': fecha_hasta
                },
                'agentes': agentes_reporte,
                'totales': {
                    'total_agentes': len(agentes_reporte),
                    'total_horas_programadas': sum(a['horas_programadas'] for a in agentes_reporte),
                    'total_horas_efectivas': sum(a['horas_efectivas'] for a in agentes_reporte),
                    'total_guardias_fines': sum(a['guardias_fines_feriados'] for a in agentes_reporte)
                }
            })

        except Exception as e:
            logger.error(f"Error generando reporte horas trabajadas: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def reporte_parte_diario(self, request):
        """Reporte de Parte Diario/Mensual Consolidado"""
        area_id = data_in.get('area')
        fecha_desde = data_in.get('fecha_desde')
        fecha_hasta = data_in.get('fecha_hasta')

        if not all([fecha_desde, fecha_hasta]):
            return Response(
                {'error': 'Se requieren parámetros: fecha_desde, fecha_hasta'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from asistencia.models import Asistencia
            from datetime import datetime, timedelta

            fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()

            # Obtener asistencias del período
            asistencias_filter = {
                'fecha__range': [fecha_inicio, fecha_fin]
            }

            if area_id:
                asistencias_filter['id_agente__area_id'] = area_id

            asistencias = Asistencia.objects.filter(
                **asistencias_filter).order_by('fecha', 'id_agente__apellido')

            # Formatear datos
            registros = []
            for asistencia in asistencias:
                # Calcular horas trabajadas
                horas_trabajadas = "N/A"
                if asistencia.hora_ingreso and asistencia.hora_egreso:
                    inicio = datetime.combine(
                        asistencia.fecha, asistencia.hora_ingreso)
                    fin = datetime.combine(
                        asistencia.fecha, asistencia.hora_egreso)
                    if fin < inicio:
                        fin += timedelta(days=1)
                    horas = (fin - inicio).total_seconds() / 3600
                    horas_trabajadas = f"{int(horas)}h {int((horas % 1) * 60)}m"

                # Determinar novedades
                novedad = "Jornada habitual"
                if asistencia.llegada_tarde:
                    novedad = "Llegada tarde"
                elif asistencia.retiro_temprano:
                    novedad = "Retiro temprano"
                elif asistencia.comision_oficial:
                    novedad = "Comisión oficial"

                registros.append({
                    'fecha': asistencia.fecha.strftime('%d/%m/%Y'),
                    'agente': f"{asistencia.id_agente.apellido}, {asistencia.id_agente.nombre}",
                    'legajo': asistencia.id_agente.legajo,
                    'hora_ingreso': asistencia.hora_ingreso.strftime('%H:%M') if asistencia.hora_ingreso else "Sin registro",
                    'hora_egreso': asistencia.hora_egreso.strftime('%H:%M') if asistencia.hora_egreso else "Sin registro",
                    'horas_trabajadas': horas_trabajadas,
                    'novedad': novedad
                })

            return Response({
                'area_nombre': 'Todas las áreas' if not area_id else 'Área seleccionada',
                'periodo': {
                    'fecha_desde': fecha_desde,
                    'fecha_hasta': fecha_hasta
                },
                'registros': registros,
                'totales': {
                    'total_registros': len(registros),
                    'total_llegadas_tarde': len([r for r in registros if 'tarde' in r['novedad']]),
                    'total_retiros_temprano': len([r for r in registros if 'temprano' in r['novedad']]),
                    'total_comisiones': len([r for r in registros if 'Comisión' in r['novedad']])
                }
            })

        except Exception as e:
            logger.error(f"Error generando parte diario: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def reporte_calculo_plus(self, request):
        """Reporte de Cálculo Plus por Guardias (20% / 40%)"""
        area_id = data_in.get('area')
        fecha_desde = data_in.get('fecha_desde')
        fecha_hasta = data_in.get('fecha_hasta')

        if not all([fecha_desde, fecha_hasta]):
            return Response(
                {'error': 'Se requieren parámetros: fecha_desde, fecha_hasta'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from datetime import datetime

            fecha_inicio = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()

            # Usar el calculador de plus existente
            from .utils import CalculadoraPlus
            calculadora = CalculadoraPlus()

            # Obtener cálculos de plus para el período
            filtros = {'fecha_desde': fecha_inicio, 'fecha_hasta': fecha_fin}
            if area_id:
                filtros['area_id'] = area_id

            resultados = calculadora.calcular_plus_periodo(**filtros)

            return Response({
                'periodo': {
                    'fecha_desde': fecha_desde,
                    'fecha_hasta': fecha_hasta
                },
                'agentes': resultados.get('agentes', []),
                'totales': {
                    'agentes_40_plus': len([a for a in resultados.get('agentes', []) if a.get('tipo_plus') == '40%']),
                    'agentes_20_plus': len([a for a in resultados.get('agentes', []) if a.get('tipo_plus') == '20%']),
                    'total_agentes_plus': len(resultados.get('agentes', []))
                }
            })

        except Exception as e:
            logger.error(f"Error generando reporte plus: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def reporte_resumen_licencias(self, request):
        """Reporte de Resumen de Licencias por agente"""
        from datetime import datetime
        area_id = data_in.get('area')

        try:
            from personas.models import Agente

            # Datos de ejemplo para licencias
            agentes_reporte = [
                {
                    'agente': 'Águila, Tayra',
                    'legajo': '001',
                    'licencia_anual_usada': 15,
                    'licencia_anual_total': 21,
                    'licencia_enfermedad_usada': 3,
                    'licencia_enfermedad_total': 30,
                    'licencia_especial_usada': 2,
                    'licencia_especial_total': 10,
                    'dias_utilizados': 20,
                    'dias_disponibles': 41
                },
                {
                    'agente': 'Alvarado, Micaela',
                    'legajo': '002',
                    'licencia_anual_usada': 8,
                    'licencia_anual_total': 21,
                    'licencia_enfermedad_usada': 0,
                    'licencia_enfermedad_total': 30,
                    'licencia_especial_usada': 1,
                    'licencia_especial_total': 10,
                    'dias_utilizados': 9,
                    'dias_disponibles': 52
                }
            ]

            return Response({
                'anio': datetime.now().year,
                'area_nombre': 'Todas las áreas' if not area_id else 'Área seleccionada',
                'agentes': agentes_reporte,
                'totales': {
                    'total_agentes': len(agentes_reporte),
                    'promedio_dias_usados': 14.5
                }
            })

        except Exception as e:
            logger.error(f"Error generando reporte licencias: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get', 'post'])
    def reporte_incumplimiento_normativo(self, request):
        """Reporte de Incumplimiento Normativo"""
        fecha_desde = data_in.get('fecha_desde')
        fecha_hasta = data_in.get('fecha_hasta')

        if not all([fecha_desde, fecha_hasta]):
            return Response(
                {'error': 'Se requieren parámetros: fecha_desde, fecha_hasta'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Datos de ejemplo
            alertas = [
                {
                    'tipo': 'exceso_horas_semanales',
                    'nivel': 'crítico',
                    'agente': 'Águila, Tayra',
                    'descripcion': 'Exceso de horas semanales: 52.0h (máximo: 48h)'
                },
                {
                    'tipo': 'descanso_insuficiente',
                    'nivel': 'advertencia',
                    'agente': 'Rodríguez, Carlos',
                    'descripcion': 'Posible descanso insuficiente entre guardias'
                }
            ]

            return Response({
                'periodo': {
                    'fecha_desde': fecha_desde,
                    'fecha_hasta': fecha_hasta
                },
                'alertas': alertas,
                'resumen': {
                    'total_alertas': 2,
                    'alertas_críticas': 1,
                    'alertas_advertencia': 1,
                    'agentes_afectados': 2
                }
            })

        except Exception as e:
            logger.error(f"Error generando reporte incumplimiento: {e}")
            return Response(
                {'error': f'Error generando reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['patch'])
    def aprobar(self, request, pk=None):
        """Aprueba una compensación individual"""
        try:
            from personas.models import Agente

            compensacion = self.get_object()

            if compensacion.estado != 'pendiente':
                return Response(
                    {'error': 'Solo se pueden aprobar compensaciones pendientes'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Obtener agente aprobador
            agente_id = request.data.get('aprobado_por')
            if not agente_id:
                return Response(
                    {'error': 'Se requiere aprobado_por'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            agente_aprobador = Agente.objects.get(id_agente=agente_id)
            observaciones = request.data.get('observaciones', '')

            # Aprobar compensación
            compensacion.aprobar(agente_aprobador, observaciones)

            # Registrar en auditoría
            Auditoria.objects.create(
                pk_afectada=compensacion.id_hora_compensacion,
                nombre_tabla='hora_compensacion',
                creado_en=timezone.now(),
                valor_previo={'estado': 'pendiente'},
                valor_nuevo={'estado': 'aprobada'},
                accion='APROBAR_COMPENSACION',
                id_agente_id=agente_aprobador.id_agente
            )

            serializer = HoraCompensacionSerializer(compensacion)
            return Response({
                'mensaje': 'Compensación aprobada exitosamente',
                'compensacion': serializer.data
            })

        except Exception as e:
            logger.error(f"Error aprobando compensación: {e}")
            return Response(
                {'error': f'Error aprobando compensación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['patch'])
    def rechazar(self, request, pk=None):
        """Rechaza una compensación individual"""
        try:
            from personas.models import Agente

            compensacion = self.get_object()

            if compensacion.estado != 'pendiente':
                return Response(
                    {'error': 'Solo se pueden rechazar compensaciones pendientes'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Obtener agente que rechaza
            agente_id = request.data.get('rechazado_por')
            if not agente_id:
                return Response(
                    {'error': 'Se requiere rechazado_por'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            agente_rechazador = Agente.objects.get(id_agente=agente_id)
            motivo_rechazo = request.data.get('motivo_rechazo', '')

            if not motivo_rechazo:
                return Response(
                    {'error': 'Se requiere motivo_rechazo'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Rechazar compensación
            compensacion.rechazar(agente_rechazador, motivo_rechazo)

            # Registrar en auditoría
            Auditoria.objects.create(
                pk_afectada=compensacion.id_hora_compensacion,
                nombre_tabla='hora_compensacion',
                creado_en=timezone.now(),
                valor_previo={'estado': 'pendiente'},
                valor_nuevo={'estado': 'rechazada'},
                accion='RECHAZAR_COMPENSACION',
                id_agente_id=agente_rechazador.id_agente
            )

            serializer = HoraCompensacionSerializer(compensacion)
            return Response({
                'mensaje': 'Compensación rechazada exitosamente',
                'compensacion': serializer.data
            })

        except Exception as e:
            logger.error(f"Error rechazando compensación: {e}")
            return Response(
                {'error': f'Error rechazando compensación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
